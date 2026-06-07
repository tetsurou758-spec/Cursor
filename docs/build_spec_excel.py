"""
build_spec_excel.py
機能概要・画面別機能仕様書をスクリーンショット付きExcelで生成する。

使い方:
    python docs/build_spec_excel.py
出力:
    docs/AX損保代理店システム_機能仕様書.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
SS_DIR    = os.path.join(BASE_DIR, "screenshots")
OUT_FILE  = os.path.join(BASE_DIR, "AX損保代理店システム_機能仕様書.xlsx")

# =========================================================
# カラーテーマ
# =========================================================
C_NAVY     = "1A2D5A"   # ヘッダー紺
C_PINK     = "FF1493"   # 社員ピンク
C_LBLUE    = "D6E4F0"   # 薄青（代理店行）
C_LPINK    = "FFD6E8"   # 薄ピンク（社員行）
C_HEADER_F = "FFFFFF"   # ヘッダー文字白
C_GRAY     = "F5F5F5"   # 奇数行背景
C_DGRAY    = "CCCCCC"   # 罫線グレー

def thin_border():
    s = Side(style="thin", color=C_DGRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(ws, row, col, value, bg=C_NAVY, fg=C_HEADER_F, bold=True, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = PatternFill("solid", fgColor=bg)
    cell.font   = Font(bold=bold, color=fg, size=size, name="Meiryo UI")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    return cell

def data_style(ws, row, col, value, bg="FFFFFF", align="left", wrap=True, bold=False, size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = PatternFill("solid", fgColor=bg)
    cell.font   = Font(bold=bold, color="333333", size=size, name="Meiryo UI")
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = thin_border()
    return cell

def merge_header(ws, row, c1, c2, value, bg=C_NAVY, fg=C_HEADER_F, size=11):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = header_style(ws, row, c1, value, bg=bg, fg=fg, size=size)
    return cell

# =========================================================
# シート1: 表紙
# =========================================================
def build_cover(wb):
    ws = wb.active
    ws.title = "表紙"
    ws.sheet_view.showGridLines = False

    # 列幅
    for i, w in enumerate([3, 15, 40, 20, 20, 3], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # タイトルブロック
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[5].height = 60
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 30

    ws.merge_cells("B5:E5")
    c = ws["B5"]
    c.value = "AX損害保険株式会社"
    c.font  = Font(bold=True, size=24, color=C_NAVY, name="Meiryo UI")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B6:E6")
    c = ws["B6"]
    c.value = "代理店Webシステム　機能仕様書"
    c.font  = Font(bold=True, size=20, color=C_NAVY, name="Meiryo UI")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B7:E7")
    c = ws["B7"]
    c.value = "Pistols Project  -  Demo Environment"
    c.font  = Font(size=12, color="888888", name="Meiryo UI")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # 仕切り線
    for col in range(2, 6):
        cell = ws.cell(row=8, column=col)
        cell.fill = PatternFill("solid", fgColor=C_NAVY)
        ws.row_dimensions[8].height = 4

    # ドキュメント情報テーブル
    info = [
        ("プロジェクト名", "Pistols  全国損害保険代理店向けWebシステム（デモ環境）"),
        ("バージョン",     "1.0.0"),
        ("作成日",         "2026-06-08"),
        ("作成者",         "開発チーム（Claude Sonnet 4.6 支援）"),
        ("システム構成",   "FastAPI (Python) + SQLite + HTML/CSS/JS"),
        ("対象ポート",     "8000"),
    ]
    ws.row_dimensions[9].height = 8
    for i, (k, v) in enumerate(info, 10):
        ws.row_dimensions[i].height = 24
        data_style(ws, i, 2, k, bg=C_LBLUE, bold=True)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
        data_style(ws, i, 3, v)

    # ログイン情報
    ws.row_dimensions[17].height = 12
    ws.row_dimensions[18].height = 24
    ws.merge_cells("B18:E18")
    header_style(ws, 18, 2, "デモアカウント一覧", bg=C_NAVY)
    ws.merge_cells("C18:E18")

    logins = [
        ("代理店ログイン", "A001", "admin",  "password123", "管理者"),
        ("代理店ログイン", "A001", "staff1", "pass001",     "一般担当"),
        ("代理店ログイン", "B002", "agent1", "pass456",     "一般担当"),
        ("代理店ログイン", "C003", "user1",  "pass789",     "閲覧専用"),
        ("社員ログイン",   "S001", "—",      "staff123",    "システム管理者"),
        ("社員ログイン",   "S002", "—",      "staff456",    "代理店担当者"),
        ("社員ログイン",   "S003", "—",      "staff789",    "参照専用"),
    ]
    hdrs = ["種別", "代理店CD/社員CD", "ログインID", "パスワード", "ロール"]
    ws.row_dimensions[19].height = 22
    for ci, h in enumerate(hdrs, 2):
        header_style(ws, 19, ci, h, bg="2E4057")
    for ri, row_data in enumerate(logins, 20):
        ws.row_dimensions[ri].height = 20
        bg = C_LPINK if "社員" in row_data[0] else C_LBLUE
        for ci, val in enumerate(row_data, 2):
            data_style(ws, ri, ci, val, bg=bg)

# =========================================================
# シート2: 機能概要
# =========================================================
def build_overview(wb):
    ws = wb.create_sheet("機能概要")
    ws.sheet_view.showGridLines = False

    col_widths = [3, 6, 22, 45, 12, 12, 3]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 8
    ws.merge_cells("B2:F2")
    header_style(ws, 2, 2, "機能概要一覧", bg=C_NAVY, size=14)
    ws.row_dimensions[2].height = 36

    hdrs = ["No.", "機能名", "概要", "社員", "代理店"]
    ws.row_dimensions[3].height = 24
    for ci, h in enumerate(hdrs, 2):
        header_style(ws, 3, ci, h)

    FEATURES = [
        (1,  "JWT認証・ブルートフォース対策",    "JWT（JSON Web Token）によるステートレス認証。ログイン失敗5回でアカウントロック（30分自動解除）。セッションタイムアウト30分。", "✓", "✓"),
        (2,  "デュアルテーマ UI",               "代理店：紺色テーマ（#1A2D5A）\n社員：ホットピンクテーマ（#FF1493）\ndata-theme属性でCSS変数を切替。", "✓", "✓"),
        (3,  "タブ別セッション管理",             "sessionStorageを使用しタブ毎に独立したセッションを管理。同一ブラウザで代理店・社員を同時ログイン可能。", "✓", "✓"),
        (4,  "RBACロール管理",                  "代理店：管理者／一般担当／閲覧専用\n社員：システム管理者／代理店担当者／参照専用\nロール×機能のマトリクスで権限制御。", "✓", "✓"),
        (5,  "ダッシュボード",                   "契約件数・更改率・保険種目別ドーナツグラフ・カレンダー・ウィジェット表示。ダイレクト検索（顧客名/TEL/証券番号自動判定）。", "✓", "✓"),
        (6,  "満期管理",                         "アコーディオン形式で月別に満期契約を一覧表示。フォローコール状況・更改STS・意向STSをインライン編集。ソート・絞込対応。", "✓", "✓"),
        (7,  "顧客管理",                         "顧客名寄せ（参照Gコード単位）。7種目の加入状況バッジ表示。未加入絞込・満期管理リンク遷移。", "✓", "✓"),
        (8,  "契約照会・契約詳細",               "代理店管轄の全契約を一覧表示。保険種目別タブ切替。契約詳細でカバレッジ・特約・車両情報等を表示。", "✓", "✓"),
        (9,  "保険金支払状況・詳細",             "事故登録・支払状況管理（未払・一部支払・支払済）。支払明細（支払種別・金額・日付）。", "✓", "✓"),
        (10, "意向確認",                         "意向確認一覧・詳細入力（顧客ニーズ・提案商品・比較商品・推奨理由・最終確定商品・顧客確認）。STS管理。", "✓", "✓"),
        (11, "コンタクト履歴",                   "顧客へのコンタクト記録（日時・手段・メモ）。代理店・社員区別なく入力可能。", "✓", "✓"),
        (12, "TODOタスク管理",                   "代理店・社員が使用できるタスク管理。期限・ステータス・メモ管理。", "✓", "✓"),
        (13, "AIレコメンド",                     "OpenAI API連携。顧客属性・契約情報から未加入種目の推奨メッセージ生成。バルク分析対応。", "✓", "✓"),
        (14, "更改おすすめプラン通知書PDF",      "ReportLab + svglib で通知書PDFを生成。DBにBLOB保存。iframeモーダル表示・ダウンロード対応。", "✓", "✓"),
        (15, "帳票管理",                         "非同期帳票生成リクエスト管理。ステータス管理（待機中/処理中/完了）。ファイルダウンロード。", "✓", "✓"),
        (16, "成績管理",                         "月別×保険種目別の更改実績集計。進捗率グラフ（目標対比）。", "—",  "✓"),
        (17, "目標設定",                         "代理店別・月別・種目別の更改目標金額設定。", "—",  "✓"),
        (18, "手数料管理",                       "保険会社別・月別の手数料明細（初年度・更改・その他）。承認ワークフロー。", "—",  "✓"),
        (19, "代理店マスタ編集",                 "社員専用。代理店情報（代理店CD・名称・住所・TEL・メール・参照Gコード・部課コード）の追加・編集・削除。", "✓",  "—"),
        (20, "権限管理",                         "ユーザー一覧・追加・編集・削除。ロール割当。パーミッションマトリクス表示。", "✓", "✓"),
        (21, "顧客名寄せバッチ",                 "夜間バッチ。性別・生年月日・名（first_name_raw）一致 AND（電話番号 OR 住所）でマッチング。参照Gコード単位で独立管理。", "—",  "—"),
        (22, "リスクマップ",                     "顧客・代理店のリスク状況をPDFマップで可視化。", "✓", "✓"),
    ]

    for ri, row_data in enumerate(FEATURES, 4):
        no, name, desc, staff, agency = row_data
        ws.row_dimensions[ri].height = 48
        bg = C_GRAY if no % 2 == 0 else "FFFFFF"
        data_style(ws, ri, 2, no,     bg=bg, align="center")
        data_style(ws, ri, 3, name,   bg=bg, bold=True)
        data_style(ws, ri, 4, desc,   bg=bg)
        c_staff  = "✓" if staff  == "✓" else "—"
        c_agency = "✓" if agency == "✓" else "—"
        data_style(ws, ri, 5, c_staff,  bg=C_LPINK if staff  == "✓" else bg, align="center", bold=(staff=="✓"))
        data_style(ws, ri, 6, c_agency, bg=C_LBLUE if agency == "✓" else bg, align="center", bold=(agency=="✓"))

# =========================================================
# 画面別仕様データ
# =========================================================
SCREEN_SPECS = [
    {
        "png":    "01_login.png",
        "title":  "代理店ログイン",
        "file":   "frontend/login.html",
        "theme":  "紺色（#1A2D5A）",
        "access": "全ユーザー（認証不要）",
        "summary": "代理店スタッフ向けのログイン画面。代理店コード・ログインID・パスワードの3項目で認証。",
        "functions": [
            ("ログイン認証",     "代理店コード・ログインID・パスワードを入力してJWTトークンを取得。sessionStorageに保存。"),
            ("ブルートフォース対策", "5回連続失敗でアカウントロック。ロック中は残り時間をカウントダウン表示。"),
            ("社員ログインリンク", "社員専用ログイン（staff_login.html）へのリンクを画面下部に表示。"),
        ],
        "api": [("POST /api/login", "代理店認証・JWT発行")],
    },
    {
        "png":    "02_staff_login.png",
        "title":  "社員ログイン",
        "file":   "frontend/staff_login.html",
        "theme":  "ホットピンク（#FF1493）",
        "access": "社員ユーザー（認証不要）",
        "summary": "AX損保社員専用のログイン画面。社員コード・パスワードの2項目で認証。ピンクテーマを適用。",
        "functions": [
            ("社員認証",         "社員コード・パスワードでJWT認証。staff_token としてsessionStorageに保存。"),
            ("ブルートフォース対策", "5回連続失敗でアカウントロック。"),
            ("代理店ログインリンク", "代理店ログイン（login.html）へのリンクを表示。"),
        ],
        "api": [("POST /api/staff/login", "社員認証・JWT発行")],
    },
    {
        "png":    "03_dashboard_agency.png",
        "title":  "ダッシュボード（代理店）",
        "file":   "frontend/dashboard.html",
        "theme":  "紺色（代理店）/ ピンク（社員）共通",
        "access": "ログイン済み全ユーザー",
        "summary": "ログイン後のトップ画面。契約・顧客状況をグラフとウィジェットで可視化。ダイレクト検索でワンアクション遷移。",
        "functions": [
            ("契約サマリーウィジェット", "総契約件数・今月満期・フォローコール未了・更改未済の件数を表示。"),
            ("保険種目別ドーナツグラフ", "7種目の契約件数比率をドーナツグラフで表示。クリックで未対応絞込遷移。"),
            ("月別更改実績グラフ",       "直近12ヶ月の更改実績（件数・保険料）を棒グラフで表示。"),
            ("カレンダーウィジェット",   "今月の満期・フォローコール予定をカレンダー表示。"),
            ("ダイレクト検索",           "顧客名/電話番号/証券番号を自動判定して顧客管理・契約照会へ遷移。"),
            ("ナビゲーション",           "全機能への画面遷移リンク（Font Awesome アイコン付き）。"),
        ],
        "api": [
            ("GET /api/dashboard/summary", "契約サマリー集計"),
            ("GET /api/dashboard/chart",   "月別実績データ"),
        ],
    },
    {
        "png":    "05_maturity.png",
        "title":  "満期管理",
        "file":   "frontend/maturity.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "満期を迎える契約を月別アコーディオンで一覧表示。フォローコール状況・更改STS・意向STSをインライン編集できる。",
        "functions": [
            ("月別アコーディオン表示", "満期月ごとに折りたたみ表示。件数バッジで一目確認。"),
            ("フォローコールSTS編集", "未コール／コール済／不要 をドロップダウンでインライン更新。"),
            ("更改STS編集",           "更改中／更改済／失効 をインライン更新。"),
            ("意向STS表示",           "意向確認登録状況（未記録／入力済／完了）をバッジ表示。"),
            ("ソート機能",            "保険料・満期日・担当者でのソート（昇順/降順）。"),
            ("絞込フィルター",        "フォローコール未了・更改未済・保険種目での絞込。"),
            ("顧客管理リンク遷移",    "顧客名クリックで顧客管理画面へ遷移（linked_customer_id経由）。"),
            ("更改おすすめ通知書PDF", "証券ごとにPDF生成・iframeモーダル表示・ダウンロード。"),
        ],
        "api": [
            ("GET /api/maturity",           "満期契約一覧取得"),
            ("PUT /api/maturity/{id}",      "STS更新"),
            ("GET /api/intentions/{ac}",    "意向確認STS取得"),
            ("GET /api/renewal/{contract_no}", "更改通知書PDF取得"),
        ],
    },
    {
        "png":    "06_customer.png",
        "title":  "顧客管理",
        "file":   "frontend/customer.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "名寄せ処理済みの顧客マスタ一覧。7種目の加入状況を視覚的に表示し、未加入絞込でアップセル対象の特定が可能。",
        "functions": [
            ("顧客一覧表示",         "参照Gコード単位でマッチングされた顧客を一覧表示。"),
            ("7種目加入状況バッジ",  "AUTO/FIRE/INJURY/JIBAI/LIABILITY/CYBER/INCOMEの加入有無をカラーバッジで表示。"),
            ("未加入絞込",           "特定種目に未加入の顧客を絞込。アップセル提案リストとして活用。"),
            ("顧客詳細表示",         "基本情報・家族構成・趣味・資産情報・コンタクト履歴を表示。"),
            ("満期管理リンク遷移",   "顧客に紐づく契約の満期一覧へワンクリック遷移。"),
            ("AIレコメンド連携",     "顧客単位のAIレコメンド結果を表示・生成。"),
            ("フリーワード検索",     "顧客名・電話番号・住所での絞込検索。"),
        ],
        "api": [
            ("GET /api/customers",           "顧客一覧取得"),
            ("GET /api/customers/{id}",      "顧客詳細取得"),
            ("GET /api/customers/{id}/contracts", "顧客別契約一覧"),
        ],
    },
    {
        "png":    "07_contract.png",
        "title":  "契約照会",
        "file":   "frontend/contract.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "代理店管轄の全契約を一覧表示。保険種目別タブ・状態別絞込・詳細遷移が可能。",
        "functions": [
            ("保険種目別タブ",     "7種目タブ切替（ALL含む）。種目ごとの契約件数バッジ表示。"),
            ("契約一覧表示",       "証券番号・顧客名・保険料・満期日・担当者・STSを一覧表示。"),
            ("更改STS絞込",        "更改中／更改済／失効での絞込フィルター。"),
            ("契約詳細遷移",       "証券番号クリックで契約詳細画面へ遷移。"),
            ("フリーワード検索",   "顧客名・証券番号での絞込検索。"),
        ],
        "api": [
            ("GET /api/contracts", "契約一覧取得（種目・STS・キーワードフィルター）"),
        ],
    },
    {
        "png":    "08_contract_detail.png",
        "title":  "契約詳細",
        "file":   "frontend/contract_detail.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "1件の契約に関するすべての情報（基本情報・カバレッジ・特約・車両/物件情報等）を表示。",
        "functions": [
            ("基本情報表示",     "証券番号・契約者情報・満期日・年間保険料・担当者・STS。"),
            ("カバレッジ表示",   "保険種目別の補償内容・保険金額・免責金額を表示。"),
            ("車両情報",         "自動車保険の場合は車名・型式・ナンバー・使用目的・等級を表示。"),
            ("物件情報",         "火災保険の場合は建物金額・家財金額・構造・地震保険フラグを表示。"),
            ("事故履歴リンク",   "この契約に紐づく事故情報へのリンク。"),
            ("更改おすすめPDF", "更改おすすめプラン通知書のiframeモーダル表示。"),
        ],
        "api": [
            ("GET /api/contracts/{id}",        "契約詳細取得"),
            ("GET /api/contract_details/{id}", "カバレッジ詳細取得"),
        ],
    },
    {
        "png":    "09_claim.png",
        "title":  "保険金支払状況",
        "file":   "frontend/claim.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "代理店管轄の事故・保険金支払情報を一覧表示。支払状況別の絞込と詳細遷移が可能。",
        "functions": [
            ("事故一覧表示",     "事故番号・発生日・報告日・事故種別・説明・支払状況を一覧表示。"),
            ("支払状況絞込",     "未払／一部支払／支払済での絞込フィルター。"),
            ("事故詳細遷移",     "事故IDクリックで保険金詳細画面へ遷移。"),
        ],
        "api": [
            ("GET /api/accidents", "事故一覧取得"),
        ],
    },
    {
        "png":    "10_claim_detail.png",
        "title":  "保険金詳細",
        "file":   "frontend/claim_detail.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "1件の事故に関する詳細情報と支払明細を表示。",
        "functions": [
            ("事故詳細表示",   "事故概要（発生日・種別・説明・STS）を表示。"),
            ("支払明細一覧",   "支払種別・支払金額・支払状況・支払日付を明細表示。"),
            ("支払合計",       "明細の支払済合計額を自動集計表示。"),
        ],
        "api": [
            ("GET /api/accidents/{id}",          "事故詳細取得"),
            ("GET /api/accidents/{id}/payments", "支払明細取得"),
        ],
    },
    {
        "png":    "11_contact.png",
        "title":  "コンタクト履歴",
        "file":   "frontend/contact.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "顧客へのコンタクト（電話・メール・訪問等）の記録を入力・参照する画面。",
        "functions": [
            ("コンタクト記録入力", "日時・コンタクト手段（TEL/メール/訪問/その他）・メモを入力して登録。"),
            ("コンタクト履歴一覧", "過去のコンタクト記録を新しい順に一覧表示。"),
            ("顧客検索",           "顧客名・代理店コードで対象顧客を検索・選択。"),
        ],
        "api": [
            ("POST /api/contacts",      "コンタクト記録登録"),
            ("GET  /api/contacts/{id}", "コンタクト履歴取得"),
        ],
    },
    {
        "png":    "12_intention.png",
        "title":  "意向確認一覧",
        "file":   "frontend/intention.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "保険募集時の意向確認情報を一覧管理。STS別絞込・ページネーション対応。",
        "functions": [
            ("意向確認一覧表示", "証券番号・顧客名・保険種目・意向STS・担当者・更新日を一覧表示。"),
            ("STS絞込",          "未記録／入力済／完了でのフィルター。"),
            ("詳細遷移",         "証券番号クリックで意向確認詳細画面へ遷移。"),
            ("ページネーション", "1ページ20件表示。前後ページナビゲーション。"),
        ],
        "api": [
            ("GET /api/intentions/{agency_code}", "意向確認一覧取得"),
        ],
    },
    {
        "png":    "13_intention_detail.png",
        "title":  "意向確認詳細",
        "file":   "frontend/intention_detail.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "1件の意向確認情報を詳細入力・更新する画面。金融庁ガイドライン対応の入力項目。",
        "functions": [
            ("意向確認入力",     "顧客ニーズ・提案商品・比較商品・推奨理由・最終確定商品・顧客確認日時を入力。"),
            ("失効理由入力",     "失効の場合は失効理由・詳細を入力。"),
            ("STS更新",          "入力済→完了のステータス遷移。"),
            ("保存トースト通知", "保存成功/失敗をページ上部固定のトースト通知で表示（3秒自動消去）。"),
        ],
        "api": [
            ("GET /api/intentions/{ac}/{policy_no}", "意向確認詳細取得"),
            ("PUT /api/intentions/{ac}/{policy_no}", "意向確認更新"),
        ],
    },
    {
        "png":    "14_admin.png",
        "title":  "権限管理",
        "file":   "frontend/admin.html",
        "theme":  "共通",
        "access": "管理者ロール以上",
        "summary": "ユーザーアカウントの作成・編集・削除とロール割当を管理する画面。パーミッションマトリクスも参照可能。",
        "functions": [
            ("ユーザー一覧",         "代理店内ユーザー（代理店スタッフ）の一覧表示。ロール・状態・最終ログイン。"),
            ("ユーザー追加モーダル", "ログインID・パスワード・氏名・ロールを設定して追加。"),
            ("ユーザー編集モーダル", "ロール変更・パスワードリセット・有効/無効切替。"),
            ("ユーザー削除",         "確認ダイアログ付きで削除。"),
            ("パーミッションマトリクス", "ロール別の機能アクセス権限を表形式で一覧表示（参照のみ）。"),
        ],
        "api": [
            ("GET    /api/users",           "ユーザー一覧取得"),
            ("POST   /api/users",           "ユーザー追加"),
            ("PUT    /api/users/{id}",      "ユーザー更新"),
            ("DELETE /api/users/{id}",      "ユーザー削除"),
            ("GET    /api/roles",           "ロール一覧取得"),
            ("GET    /api/permissions",     "パーミッション取得"),
        ],
    },
    {
        "png":    "15_agency_master.png",
        "title":  "代理店マスタ編集",
        "file":   "frontend/agency_master.html",
        "theme":  "ホットピンク（社員専用）",
        "access": "社員ログイン必須",
        "summary": "社員専用の代理店マスタ編集画面。代理店情報の追加・編集・削除・検索が可能。",
        "functions": [
            ("代理店一覧表示",   "全代理店の一覧表示（代理店CD・名称・住所・TEL・参照Gコード・部課コード）。"),
            ("代理店検索",       "代理店CD・名称・部課コードでの絞込検索。"),
            ("代理店追加モーダル", "新規代理店情報の入力・登録。"),
            ("代理店編集モーダル", "既存代理店情報の編集・保存。"),
            ("代理店削除",       "確認ダイアログ付きで削除。代理店ユーザーも連動削除。"),
        ],
        "api": [
            ("GET    /api/agencies",       "代理店一覧取得"),
            ("POST   /api/agencies",       "代理店追加"),
            ("PUT    /api/agencies/{id}",  "代理店更新"),
            ("DELETE /api/agencies/{id}", "代理店削除"),
        ],
    },
    {
        "png":    "16_sales.png",
        "title":  "成績管理",
        "file":   "frontend/sales.html",
        "theme":  "紺色（代理店）",
        "access": "代理店ログイン（管理者・一般担当）",
        "summary": "月別×保険種目別の更改実績を集計・グラフ表示。目標対比の進捗率を可視化。",
        "functions": [
            ("月別×種目別実績グラフ", "7種目×12ヶ月の更改実績（件数・保険料）を棒グラフで表示。"),
            ("進捗率表示",           "月別目標対比の達成率をパーセントで表示。"),
            ("担当者別フィルター",   "担当者別での実績絞込。"),
            ("目標設定画面リンク",   "目標設定画面へのショートカット。"),
        ],
        "api": [
            ("GET /api/sales/monthly",  "月別実績集計"),
            ("GET /api/sales/targets",  "目標値取得"),
        ],
    },
    {
        "png":    "17_sales_target.png",
        "title":  "目標設定",
        "file":   "frontend/sales_target.html",
        "theme":  "紺色（代理店）",
        "access": "代理店ログイン（管理者）",
        "summary": "代理店の月別×保険種目別の更改目標金額を設定する画面。",
        "functions": [
            ("目標金額入力",     "種目別・月別の目標金額をグリッド形式で一括入力。"),
            ("年度切替",         "年度プルダウンで切替。"),
            ("保存",             "変更した目標値を一括保存。"),
        ],
        "api": [
            ("GET /api/sales/targets",           "目標値取得"),
            ("POST /api/sales/targets",          "目標値設定"),
        ],
    },
    {
        "png":    "18_commission.png",
        "title":  "手数料管理",
        "file":   "frontend/commission.html",
        "theme":  "紺色（代理店）",
        "access": "代理店ログイン（管理者）",
        "summary": "保険会社別・月別の代理店手数料明細を管理する画面。承認ワークフロー対応。",
        "functions": [
            ("手数料明細一覧",   "保険会社別・月別の初年度手数料・更改手数料・その他手数料を一覧表示。"),
            ("承認ステータス",   "未承認／承認済／否認のワークフロー管理。"),
            ("合計集計",         "月別・種別の手数料合計を自動集計表示。"),
            ("追加・編集",       "手数料レコードの追加・編集モーダル。"),
        ],
        "api": [
            ("GET    /api/commissions",      "手数料一覧取得"),
            ("POST   /api/commissions",      "手数料追加"),
            ("PUT    /api/commissions/{id}", "手数料更新"),
        ],
    },
    {
        "png":    "19_ai_recommend.png",
        "title":  "AIレコメンド",
        "file":   "frontend/ai_recommend.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "OpenAI API連携によるAI推薦機能。顧客属性・契約情報から未加入種目の推奨メッセージを生成。バルク分析にも対応。",
        "functions": [
            ("推薦一覧表示",         "顧客別の推薦種目・推薦理由・リスクスコアを一覧表示。"),
            ("AI推薦生成",           "選択した顧客に対してOpenAI APIで推薦テキストを生成。"),
            ("バルク分析",           "複数顧客を一括でAI分析。バルクジョブIDで管理。"),
            ("社員向け代理店選択",   "社員ログイン時は管轄代理店をドロップダウンで選択。"),
            ("リスクスコア表示",     "0〜100のリスクスコアをカラーバーで視覚表示。"),
        ],
        "api": [
            ("GET  /api/ai/recommendations/{agency}", "推薦一覧取得"),
            ("POST /api/ai/recommend",                "単件AI推薦生成"),
            ("POST /api/ai/recommend/bulk",           "バルクAI推薦生成"),
        ],
    },
    {
        "png":    "20_report_list.png",
        "title":  "帳票管理",
        "file":   "frontend/report_list.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "帳票の非同期生成リクエストを管理する画面。生成ステータスの確認とダウンロードが可能。",
        "functions": [
            ("帳票リクエスト一覧", "帳票種別・リクエスト日時・ステータス（待機中/処理中/完了/エラー）・ファイルサイズを表示。"),
            ("ダウンロード",       "完了済み帳票のExcel/PDFダウンロード。"),
            ("ステータス自動更新", "処理中の帳票は30秒ごとにステータスを自動ポーリング。"),
            ("帳票生成リクエスト", "帳票種別・対象期間を指定して生成リクエストを送信。"),
        ],
        "api": [
            ("GET  /api/reports",        "帳票リクエスト一覧取得"),
            ("POST /api/reports/request","帳票生成リクエスト"),
            ("GET  /api/reports/{id}/download", "帳票ファイルダウンロード"),
        ],
    },
    {
        "png":    "21_todo_list.png",
        "title":  "TODOリスト",
        "file":   "frontend/todo_list.html",
        "theme":  "共通",
        "access": "ログイン済み全ユーザー",
        "summary": "代理店・社員が使用できるタスク管理機能。期限・ステータス・メモを管理。",
        "functions": [
            ("TODO一覧表示",   "タイトル・説明・期日・ステータス・担当者を一覧表示。"),
            ("TODO追加",       "タイトル・説明・期日・担当者を指定して追加。"),
            ("STS更新",        "未着手→進行中→完了のステータス管理。"),
            ("TODO削除",       "確認ダイアログ付きで削除。"),
            ("絞込フィルター", "ステータス・担当者での絞込。"),
        ],
        "api": [
            ("GET    /api/todos",       "TODO一覧取得"),
            ("POST   /api/todos",       "TODO追加"),
            ("PUT    /api/todos/{id}",  "TODO更新"),
            ("DELETE /api/todos/{id}", "TODO削除"),
        ],
    },
]

# =========================================================
# シート3以降: 画面別仕様
# =========================================================
def build_screen_sheet(wb, spec):
    title = spec["title"]
    ws = wb.create_sheet(title[:20])  # シート名は31文字まで
    ws.sheet_view.showGridLines = False

    # 列幅
    col_widths = [2, 18, 55, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws.row_dimensions[row].height = 8
    row += 1

    # タイトル行
    ws.row_dimensions[row].height = 36
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    header_style(ws, row, 2, f"画面仕様書：{title}", bg=C_NAVY, size=14)
    row += 1

    # 基本情報
    basic = [
        ("画面ファイル",   spec["file"]),
        ("テーマ",         spec["theme"]),
        ("アクセス権限",   spec["access"]),
        ("概要",           spec["summary"]),
    ]
    for label, val in basic:
        ws.row_dimensions[row].height = 36
        data_style(ws, row, 2, label, bg=C_LBLUE, bold=True)
        data_style(ws, row, 3, val)
        row += 1

    row += 1  # 空行

    # スクリーンショット
    png_path = os.path.join(SS_DIR, spec["png"])
    if os.path.exists(png_path):
        ws.row_dimensions[row].height = 18
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        header_style(ws, row, 2, "スクリーンショット", bg="2E4057", size=11)
        row += 1

        try:
            img = XLImage(png_path)
            # 幅を列幅に合わせてリサイズ（約1000px幅）
            MAX_W = 900
            ratio = MAX_W / img.width
            img.width  = int(img.width  * ratio)
            img.height = int(img.height * ratio)
            ws.add_image(img, f"B{row}")
            # 画像の高さ分の行を確保（1px ≈ 0.75pt ≈ 1行）
            img_rows = max(1, img.height // 15)
            for r in range(row, row + img_rows + 1):
                ws.row_dimensions[r].height = 15
            row += img_rows + 2
        except Exception as e:
            data_style(ws, row, 2, f"[スクリーンショット読込エラー: {e}]", bg="FFF0F0")
            row += 2
    else:
        ws.row_dimensions[row].height = 24
        data_style(ws, row, 2, f"[スクリーンショット未取得: {spec['png']}]", bg="FFF0F0")
        row += 2

    # 機能仕様
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    header_style(ws, row, 2, "機能仕様", bg="2E4057", size=11)
    row += 1

    ws.row_dimensions[row].height = 22
    header_style(ws, row, 2, "機能名",  bg="3A5070", size=10)
    header_style(ws, row, 3, "仕様説明", bg="3A5070", size=10)
    row += 1

    for i, (fname, fdesc) in enumerate(spec["functions"]):
        ws.row_dimensions[row].height = 42
        bg = C_GRAY if i % 2 == 0 else "FFFFFF"
        data_style(ws, row, 2, fname, bg=bg, bold=True)
        data_style(ws, row, 3, fdesc, bg=bg)
        row += 1

    row += 1

    # API一覧
    if spec.get("api"):
        ws.row_dimensions[row].height = 18
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        header_style(ws, row, 2, "使用APIエンドポイント", bg="2E4057", size=11)
        row += 1

        ws.row_dimensions[row].height = 22
        header_style(ws, row, 2, "エンドポイント", bg="3A5070", size=10)
        header_style(ws, row, 3, "説明",            bg="3A5070", size=10)
        row += 1

        for i, (ep, desc) in enumerate(spec["api"]):
            ws.row_dimensions[row].height = 22
            bg = C_GRAY if i % 2 == 0 else "FFFFFF"
            data_style(ws, row, 2, ep,   bg=bg, bold=True)
            data_style(ws, row, 3, desc, bg=bg)
            row += 1

# =========================================================
# メイン
# =========================================================
def main():
    print("仕様書Excel生成中...")
    wb = Workbook()

    print("  [1/3] 表紙シート生成...")
    build_cover(wb)

    print("  [2/3] 機能概要シート生成...")
    build_overview(wb)

    print("  [3/3] 画面別仕様シート生成...")
    for spec in SCREEN_SPECS:
        print(f"         {spec['title']}...")
        build_screen_sheet(wb, spec)

    wb.save(OUT_FILE)
    size_kb = os.path.getsize(OUT_FILE) // 1024
    print(f"\n完了 -> {OUT_FILE}  ({size_kb} KB)")

if __name__ == "__main__":
    main()
