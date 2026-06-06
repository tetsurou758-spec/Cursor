"""
report_batch.py
---------------
ディレード帳票生成バッチ

対象: status='受付中' の帳票を古い順に最大5件処理する。
実行: python batch/report_batch.py
"""

import sqlite3
import datetime
import json
import os
import sys

# プロジェクトルートをパスに追加
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

# openpyxl が使えない場合は csv にフォールバック
try:
    import openpyxl
    USE_EXCEL = True
except ImportError:
    import csv
    import io
    USE_EXCEL = False

DB_PATH = os.path.join(os.path.dirname(__file__), "../db/users.sqlite")
MAX_BATCH = 5  # 1回の実行で処理する最大件数


def get_conn() -> sqlite3.Connection:
    """SQLite接続を取得する"""
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


# ───────────────────────────────────────────────────────────────────────
# 帳票種別ごとのデータ取得・生成関数
# ───────────────────────────────────────────────────────────────────────

def build_customer_list(conn: sqlite3.Connection, params: dict) -> tuple[list, list]:
    """
    顧客一覧 XLSX を生成するためのデータ取得。
    返却: (ヘッダーリスト, データ行リスト)
    """
    header = ["顧客ID", "姓", "名", "生年月日", "性別", "住所", "電話番号", "加入種目数"]

    sql = """
        SELECT
            c.customer_id,
            c.last_name,
            c.first_name,
            c.birth_date,
            CASE c.gender WHEN 'M' THEN '男性' WHEN 'F' THEN '女性' ELSE c.gender END AS gender,
            c.address,
            c.tel,
            COUNT(DISTINCT ct.id) AS contract_count
        FROM customers c
        LEFT JOIN contracts ct ON ct.linked_customer_id = c.customer_id
    """
    wheres = []
    sql_params = []

    group_code = params.get("group_code", "")
    name       = params.get("name", "")
    if group_code:
        wheres.append("c.group_code = ?")
        sql_params.append(group_code)
    if name:
        wheres.append("(c.last_name LIKE ? OR c.first_name LIKE ?)")
        sql_params += [f"%{name}%", f"%{name}%"]

    if wheres:
        sql += " WHERE " + " AND ".join(wheres)
    sql += " GROUP BY c.customer_id ORDER BY c.customer_id"

    rows = conn.execute(sql, sql_params).fetchall()
    data = [
        [r["customer_id"], r["last_name"], r["first_name"], r["birth_date"],
         r["gender"], r["address"], r["tel"], r["contract_count"]]
        for r in rows
    ]
    return header, data


def build_maturity_list(conn: sqlite3.Connection, params: dict) -> tuple[list, list]:
    """
    満期契約一覧 XLSX を生成するためのデータ取得。
    返却: (ヘッダーリスト, データ行リスト)
    """
    header = ["証券番号", "顧客名", "種目", "満期日", "保険料", "更改STS", "担当者"]

    sql = """
        SELECT
            ct.contract_no,
            c.last_name || ' ' || c.first_name AS customer_name,
            ct.policy_type,
            ct.expiry_date,
            ct.annual_premium,
            ct.renewal_status,
            ct.staff_code
        FROM contracts ct
        LEFT JOIN customers c ON ct.linked_customer_id = c.customer_id
    """
    wheres = []
    sql_params = []

    agency_code = params.get("agency_code", "")
    month       = params.get("month", "")  # 例: 2026-07
    if agency_code:
        wheres.append("ct.agency_code = ?")
        sql_params.append(agency_code)
    if month:
        wheres.append("ct.expiry_date LIKE ?")
        sql_params.append(f"{month}%")

    if wheres:
        sql += " WHERE " + " AND ".join(wheres)
    sql += " ORDER BY ct.expiry_date"

    rows = conn.execute(sql, sql_params).fetchall()
    data = [
        [r["contract_no"], r["customer_name"], r["policy_type"],
         r["expiry_date"], r["annual_premium"], r["renewal_status"], r["staff_code"]]
        for r in rows
    ]
    return header, data


def build_sales_list(conn: sqlite3.Connection, params: dict) -> tuple[list, list]:
    """
    成績一覧 XLSX を生成するためのデータ取得。
    返却: (ヘッダーリスト, データ行リスト)
    """
    header = ["代理店コード", "種目", "年", "月", "目標件数", "実績件数", "進捗率(%)"]

    sql = """
        SELECT
            st.agency_code,
            st.policy_type_code,
            st.year,
            st.month,
            st.target_count,
            st.actual_count,
            CASE WHEN st.target_count > 0
                 THEN ROUND(st.actual_count * 100.0 / st.target_count, 1)
                 ELSE NULL
            END AS progress_rate
        FROM sales_targets st
    """
    wheres = []
    sql_params = []

    agency_code = params.get("agency_code", "")
    year        = params.get("year", "")
    if agency_code:
        wheres.append("st.agency_code = ?")
        sql_params.append(agency_code)
    if year:
        wheres.append("st.year = ?")
        sql_params.append(int(year))

    if wheres:
        sql += " WHERE " + " AND ".join(wheres)
    sql += " ORDER BY st.agency_code, st.year, st.month, st.policy_type_code"

    rows = conn.execute(sql, sql_params).fetchall()
    data = [
        [r["agency_code"], r["policy_type_code"], r["year"], r["month"],
         r["target_count"], r["actual_count"], r["progress_rate"]]
        for r in rows
    ]
    return header, data


# ───────────────────────────────────────────────────────────────────────
# XLSX / CSV 生成
# ───────────────────────────────────────────────────────────────────────

def make_xlsx(header: list, data: list) -> bytes:
    """openpyxlでXLSXバイト列を生成する"""
    wb = openpyxl.Workbook()
    ws = wb.active

    # ヘッダー行（太字・背景色）
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="003366")
    for ci, col in enumerate(header, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # データ行
    for ri, row in enumerate(data, start=2):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)

    # 列幅の自動調整（最大40文字）
    for ci, col in enumerate(header, start=1):
        max_len = max(
            len(str(col)),
            max((len(str(row[ci - 1])) for row in data), default=0)
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_len + 2, 42)

    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_csv(header: list, data: list) -> bytes:
    """openpyxl不使用時のCSVフォールバック"""
    import io as _io
    buf = _io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(data)
    return buf.getvalue().encode("utf-8-sig")


def generate_file(report_type: str, params: dict, conn: sqlite3.Connection) -> tuple[bytes, str, int]:
    """
    帳票種別に応じてファイルバイト列・ファイル名拡張子・件数を返す。
    返却: (file_bytes, extension, record_count)
    """
    now_str = datetime.datetime.now().strftime("%Y%m%d")
    if report_type == "customer_list":
        header, data = build_customer_list(conn, params)
        ext = ".xlsx" if USE_EXCEL else ".csv"
    elif report_type == "maturity_list":
        header, data = build_maturity_list(conn, params)
        ext = ".xlsx" if USE_EXCEL else ".csv"
    elif report_type == "sales_list":
        header, data = build_sales_list(conn, params)
        ext = ".xlsx" if USE_EXCEL else ".csv"
    else:
        raise ValueError(f"未知の帳票種別: {report_type}")

    if USE_EXCEL:
        file_bytes = make_xlsx(header, data)
    else:
        file_bytes = make_csv(header, data)

    return file_bytes, ext, len(data)


# ───────────────────────────────────────────────────────────────────────
# バッチメイン処理
# ───────────────────────────────────────────────────────────────────────

def process_pending(conn: sqlite3.Connection):
    """受付中の帳票を最大5件取得して処理する"""
    rows = conn.execute("""
        SELECT id, request_no, agency_code, login_id, report_type, report_name, search_params
        FROM report_requests
        WHERE status = '受付中'
        ORDER BY requested_at ASC
        LIMIT ?
    """, (MAX_BATCH,)).fetchall()

    if not rows:
        print("処理対象の帳票はありません。")
        return

    for row in rows:
        req_id     = row["id"]
        request_no = row["request_no"]
        print(f"[処理開始] {request_no} ({row['report_type']})")

        # ── 処理開始：status を「作成中」に更新
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute(
            "UPDATE report_requests SET status='作成中', started_at=? WHERE id=?",
            (now, req_id)
        )
        conn.commit()

        try:
            # 検索条件を取得
            params = {}
            if row["search_params"]:
                try:
                    params = json.loads(row["search_params"])
                except json.JSONDecodeError:
                    pass

            # ファイル生成
            file_bytes, ext, record_count = generate_file(row["report_type"], params, conn)
            file_name = f"{row['report_name'].replace(' ', '_')}{ext}"
            file_size = len(file_bytes)

            # ── 完了更新
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute("""
                UPDATE report_requests
                SET status='完了', completed_at=?, record_count=?, file_size=?,
                    file_data=?, file_name=?
                WHERE id=?
            """, (now, record_count, file_size, file_bytes, file_name, req_id))
            conn.commit()
            print(f"[完了] {request_no} → {file_name} ({record_count}件, {file_size:,}bytes)")

        except Exception as e:
            # ── 異常終了更新
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute("""
                UPDATE report_requests
                SET status='異常終了', completed_at=?, error_message=?
                WHERE id=?
            """, (now, str(e), req_id))
            conn.commit()
            print(f"[エラー] {request_no}: {e}")


def main():
    print(f"[report_batch] 開始: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    conn = get_conn()
    try:
        process_pending(conn)
    finally:
        conn.close()
    print(f"[report_batch] 終了: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
