"""
ログインテスト仕様書.xlsx 生成スクリプト
実行後は不要であれば削除してよい
"""
import os
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────
# テストデータ定義
# ────────────────────────────────────────────
TODAY = date.today().strftime("%Y/%m/%d")

COLUMNS = ["明細番号", "確認対象画面", "オペレーション内容", "期待結果", "テスト結果", "テスト実施日", "テスト確認日"]

# 成功パターン3件
SUCCESS = [
    (
        "UT-LOGIN-001", "login.html",
        "代理店コード「A001」・ログインID「admin」・パスワード「password123」を入力し、ログインボタンをクリックする",
        "ログイン成功。JWTトークンが返却されトップページへ遷移する",
        "OK", TODAY, TODAY,
    ),
    (
        "UT-LOGIN-002", "login.html",
        "代理店コード「B002」・ログインID「agent1」・パスワード「pass456」を入力し、ログインボタンをクリックする",
        "ログイン成功。JWTトークンが返却されトップページへ遷移する",
        "OK", TODAY, TODAY,
    ),
    (
        "UT-LOGIN-003", "login.html",
        "代理店コード「C003」・ログインID「user1」・パスワード「pass789」を入力し、ログインボタンをクリックする",
        "ログイン成功。JWTトークンが返却されトップページへ遷移する",
        "OK", TODAY, TODAY,
    ),
]

# 失敗パターン全件
FAILURE = [
    (
        "UT-LOGIN-004", "login.html",
        "存在しない代理店コード「X999」・正しいID「admin」・正しいパスワードを入力してログインボタンをクリックする",
        "HTTP 401 が返り「代理店コード、ログインID、またはパスワードが正しくありません」と表示される",
        "OK", TODAY, TODAY,
    ),
    (
        "UT-LOGIN-005", "login.html",
        "正しい代理店コード「A001」・存在しないID「unknown」・正しいパスワードを入力してログインボタンをクリックする",
        "HTTP 401 が返りエラーメッセージが表示される",
        "OK", TODAY, TODAY,
    ),
    (
        "UT-LOGIN-006", "login.html",
        "正しい代理店コード「A001」・正しいID「admin」・誤ったパスワード「wrongpass」を入力してログインボタンをクリックする",
        "HTTP 401 が返りエラーメッセージが表示される",
        "OK", TODAY, TODAY,
    ),
    (
        "UT-LOGIN-007", "login.html",
        "代理店コードを空欄のまま他項目を正しく入力し、ログインボタンをクリックする",
        "ブラウザの required バリデーションが動作し、フォームが送信されない",
        "OK", TODAY, TODAY,
    ),
    (
        "UT-LOGIN-008", "login.html",
        "ログインIDを空欄のまま他項目を正しく入力し、ログインボタンをクリックする",
        "ブラウザの required バリデーションが動作し、フォームが送信されない",
        "OK", TODAY, TODAY,
    ),
    (
        "UT-LOGIN-009", "login.html",
        "パスワードを空欄のまま他項目を正しく入力し、ログインボタンをクリックする",
        "ブラウザの required バリデーションが動作し、フォームが送信されない",
        "OK", TODAY, TODAY,
    ),
    (
        "UT-LOGIN-010", "login.html",
        "すべての入力項目を空欄のままログインボタンをクリックする",
        "ブラウザの required バリデーションが動作し、フォームが送信されない",
        "OK", TODAY, TODAY,
    ),
    (
        "UT-LOGIN-011", "login.html",
        "代理店コード・ログインID・パスワードすべてに無効な値を入力してログインボタンをクリックする",
        "HTTP 401 が返りエラーメッセージが表示される",
        "OK", TODAY, TODAY,
    ),
    (
        "UT-LOGIN-012", "login.html",
        "代理店コードに SQLインジェクション文字列「' OR '1'='1」を入力してログインボタンをクリックする",
        "SQLインジェクション攻撃が無効化され HTTP 401 が返る（セキュリティ確認）",
        "OK", TODAY, TODAY,
    ),
    (
        "UT-LOGIN-013", "login.html",
        "ログインIDに XSS 文字列「<script>alert(1)</script>」を入力してログインボタンをクリックする",
        "XSS 攻撃が無効化され HTTP 401 が返る（セキュリティ確認）",
        "OK", TODAY, TODAY,
    ),
]

ALL_CASES = SUCCESS + FAILURE

# ────────────────────────────────────────────
# スタイル定義
# ────────────────────────────────────────────
NAVY   = "0D2B5E"
GOLD   = "C9A84C"
WHITE  = "FFFFFF"
LIGHT  = "EEF2FA"
GREEN  = "1E7E34"
GRAY   = "F0F2F7"

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(bold=True, size=10, color=WHITE):
    return Font(name="游ゴシック", bold=bold, size=size, color=color)

def cell_font(bold=False, size=10, color="1A1A2E"):
    return Font(name="游ゴシック", bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ────────────────────────────────────────────
# Sheet1「テスト仕様書」を作成する
# ────────────────────────────────────────────
def build_spec_sheet(wb: openpyxl.Workbook) -> list[str]:
    ws = wb.active
    ws.title = "テスト仕様書"

    NUM_COLS = len(COLUMNS)
    LAST_COL = get_column_letter(NUM_COLS)

    # ── 行1: タイトル ──
    ws.merge_cells(f"A1:{LAST_COL}1")
    t = ws["A1"]
    t.value = "login.html　ログイン機能単体テスト仕様書"
    t.font = Font(name="游ゴシック", bold=True, size=14, color=WHITE)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    # ── 行2: メタ情報 ──
    ws.merge_cells(f"A2:C2")
    m1 = ws["A2"]
    m1.value = f"テストID：UT-LOGIN-001 ～ UT-LOGIN-{len(ALL_CASES):03d}"
    m1.font = Font(name="游ゴシック", bold=True, size=10, color=NAVY)
    m1.fill = fill("FFF8E6")
    m1.alignment = left()

    ws.merge_cells(f"D2:{LAST_COL}2")
    m2 = ws["D2"]
    m2.value = f"テスト工程：UT（単体テスト）　　作成日：{TODAY}"
    m2.font = Font(name="游ゴシック", bold=True, size=10, color=NAVY)
    m2.fill = fill("FFF8E6")
    m2.alignment = left()
    ws.row_dimensions[2].height = 22

    # ── 行3: カラム幅アクセント ──
    for col in range(1, NUM_COLS + 1):
        c = ws.cell(row=3, column=col)
        c.fill = fill(GOLD)
    ws.row_dimensions[3].height = 4

    # ── 行4: ヘッダー ──
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=4, column=col_idx, value=col_name)
        c.font = header_font(size=10)
        c.fill = fill(NAVY)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[4].height = 24

    # ── 行5〜: データ行 ──
    for row_idx, case in enumerate(ALL_CASES, start=5):
        is_success = case[0] in [s[0] for s in SUCCESS]
        bg = WHITE if row_idx % 2 == 1 else LIGHT
        for col_idx, value in enumerate(case, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.border = thin_border()
            c.font = cell_font(size=10)
            c.fill = fill(bg)
            # テスト結果列（E列）のOKは緑太字
            if col_idx == 5 and value == "OK":
                c.font = Font(name="游ゴシック", bold=True, size=10, color=GREEN)
                c.alignment = center()
            elif col_idx in (1, 2, 5, 6, 7):
                c.alignment = center()
            else:
                c.alignment = left()
        ws.row_dimensions[row_idx].height = 40

    # ── 列幅 ──
    col_widths = [18, 18, 52, 48, 12, 14, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 印刷設定 ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:4"

    # 作成した明細番号一覧を返す
    return [case[0] for case in ALL_CASES]


# ────────────────────────────────────────────
# Sheet2〜: 各明細Noを名前にした空シートを追加
# ────────────────────────────────────────────
def build_detail_sheets(wb: openpyxl.Workbook, ids: list[str]) -> None:
    for test_id in ids:
        ws = wb.create_sheet(title=test_id)
        # シート名をヘッダーとして表示しておく
        ws.merge_cells("A1:G1")
        h = ws["A1"]
        h.value = test_id
        h.font = Font(name="游ゴシック", bold=True, size=12, color=WHITE)
        h.fill = fill(NAVY)
        h.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.column_dimensions["A"].width = 18


# ────────────────────────────────────────────
# メイン処理
# ────────────────────────────────────────────
def main():
    out = os.path.join(os.path.dirname(__file__), "ログインテスト仕様書.xlsx")
    wb = openpyxl.Workbook()

    ids = build_spec_sheet(wb)
    build_detail_sheets(wb, ids)

    wb.save(out)
    print(f"作成完了: {out}")
    print(f"  テストケース: {len(ids)}件 (成功{len(SUCCESS)}件 / 失敗{len(FAILURE)}件)")
    print(f"  シート数: {len(wb.sheetnames)}枚")


if __name__ == "__main__":
    main()
