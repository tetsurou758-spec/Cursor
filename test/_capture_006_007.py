"""
006/007 テスト証跡 スクリーンショット取得スクリプト
- UT-NAMEYOSE-001~003 (006テスト)
- UT-CUSTOMER-001~003 (007テスト)
撮影後に各証跡Excelのシートへ自動挿入する
"""
import subprocess, time, sys, pathlib, os
from playwright.sync_api import sync_playwright

BASE_DIR   = pathlib.Path(r"C:\Users\yoshi\OneDrive\ドキュメント\Cursor")
FRONTEND   = BASE_DIR / "frontend"
SS006      = BASE_DIR / "test" / "006_UT-名寄せバッチ処理テスト" / "screenshots"
SS007      = BASE_DIR / "test" / "007_UT-顧客管理保険加入状況テスト" / "screenshots"
EXCEL006   = BASE_DIR / "test" / "006_UT-名寄せバッチ処理テスト" / "名寄せテスト証跡.xlsx"
EXCEL007   = BASE_DIR / "test" / "007_UT-顧客管理保険加入状況テスト" / "加入状況テスト証跡.xlsx"
BASE_URL   = "http://127.0.0.1:8080"

SS006.mkdir(exist_ok=True)
SS007.mkdir(exist_ok=True)

screenshots = {}  # name -> path


def ss(page, name, out_dir, full_page=False):
    path = str(out_dir / name)
    page.screenshot(path=path, full_page=full_page)
    screenshots[name] = path
    print(f"  撮影: {name}")
    return path


def login_agency(page, agency_code, login_id, password):
    page.goto(f"{BASE_URL}/login.html")
    page.wait_for_load_state("networkidle")
    page.fill("#agency-code", agency_code)
    page.fill("#login-id",    login_id)
    page.fill("#password",    password)
    page.click("#btn-login")
    page.wait_for_url(f"{BASE_URL}/dashboard.html", timeout=10000)
    page.wait_for_timeout(1500)
    print(f"  代理店ログイン完了: {agency_code}/{login_id}")


def login_staff(page, staff_code, password):
    page.goto(f"{BASE_URL}/staff_login.html")
    page.wait_for_load_state("networkidle")
    page.fill("#staff-code", staff_code)
    page.fill("#password",   password)
    page.click("#btn-login")
    page.wait_for_url(f"{BASE_URL}/dashboard.html", timeout=10000)
    page.wait_for_timeout(1500)
    print(f"  社員ログイン完了: {staff_code}")


def search_customer(page, last_name, wait_ms=2500):
    page.goto(f"{BASE_URL}/customer.html")
    page.wait_for_load_state("networkidle")
    page.wait_for_timeout(1000)
    page.fill("#q-last-name", last_name)
    page.wait_for_timeout(300)
    page.click("button.btn-search")
    page.wait_for_timeout(wait_ms)
    print(f"  検索完了: 「{last_name}」")


def run():
    srv = subprocess.Popen(
        [sys.executable, "-m", "http.server", "8080", "--directory", str(FRONTEND)],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
    )
    time.sleep(1.5)

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            vp = {"width": 1400, "height": 850}

            # ============================================================
            # Phase 1: A001/admin セッション
            # UT-NAMEYOSE-001, 003 / UT-CUSTOMER-001, 002, 003
            # ============================================================
            print("\n■ Phase 1: A001/admin セッション")
            ctx1 = browser.new_context(viewport=vp)
            page1 = ctx1.new_page()

            login_agency(page1, "A001", "admin", "password123")

            # --- UT-NAMEYOSE-001 / UT-CUSTOMER-001 ---
            # 羽生を検索 → グループA 1件確認
            print("\n  UT-NAMEYOSE-001 / UT-CUSTOMER-001: 羽生検索")
            search_customer(page1, "羽生")
            page1.wait_for_selector("#customer-tbody tr", timeout=5000)
            page1.evaluate("window.scrollTo(0, 0)")
            page1.wait_for_timeout(500)
            ss(page1, "nameyose_001_hanyu_groupA.png", SS006)
            ss(page1, "customer_001_hanyu_status.png", SS007)

            # --- UT-CUSTOMER-002 ---
            # 羽生の詳細を展開
            print("  UT-CUSTOMER-002: 羽生詳細展開")
            page1.click(".toggle-btn")
            page1.wait_for_timeout(2000)
            page1.evaluate("window.scrollTo(0, 0)")
            page1.wait_for_timeout(500)
            ss(page1, "customer_002_hanyu_detail.png", SS007)

            # --- UT-NAMEYOSE-003 ---
            # 大谷を検索 → グループA 1件・複数代理店バッジ
            print("  UT-NAMEYOSE-003: 大谷検索")
            search_customer(page1, "大谷")
            page1.wait_for_selector("#customer-tbody tr", timeout=5000)
            page1.evaluate("window.scrollTo(0, 0)")
            page1.wait_for_timeout(500)
            ss(page1, "nameyose_003_ohtani_groupA.png", SS006)

            # --- UT-CUSTOMER-003 ---
            # 満期管理 → 鈴木の顧客リンクをクリック → 顧客管理遷移
            print("  UT-CUSTOMER-003: 満期管理→顧客管理遷移")
            page1.goto(f"{BASE_URL}/maturity.html")
            page1.wait_for_load_state("networkidle")
            page1.wait_for_timeout(2500)
            # 鈴木の顧客リンクを探してクリック
            page1.wait_for_selector("a.customer-link", timeout=8000)
            # 「鈴木」を含むリンクを優先して探す
            suzuki_link = page1.locator("a.customer-link").filter(has_text="鈴木").first
            if suzuki_link.count() > 0:
                suzuki_link.click()
            else:
                page1.locator("a.customer-link").first.click()
            page1.wait_for_load_state("networkidle")
            page1.wait_for_timeout(2500)
            page1.evaluate("window.scrollTo(0, 0)")
            page1.wait_for_timeout(500)
            ss(page1, "customer_003_maturity_transition.png", SS007)

            ctx1.close()

            # ============================================================
            # Phase 2: S001/staff セッション
            # UT-NAMEYOSE-002: 全グループ参照 → 羽生2件表示
            # ============================================================
            print("\n■ Phase 2: S001/staff セッション")
            ctx2 = browser.new_context(viewport=vp)
            page2 = ctx2.new_page()

            login_staff(page2, "S001", "staff123")

            # --- UT-NAMEYOSE-002 ---
            print("  UT-NAMEYOSE-002: 羽生検索（全グループ）")
            search_customer(page2, "羽生")
            page2.wait_for_selector("#customer-tbody tr", timeout=5000)
            page2.evaluate("window.scrollTo(0, 0)")
            page2.wait_for_timeout(500)
            ss(page2, "nameyose_002_hanyu_all_groups.png", SS006)

            ctx2.close()
            browser.close()

        print(f"\n[OK] スクリーンショット取得完了")
        print(f"  006: {SS006}")
        print(f"  007: {SS007}")

    finally:
        srv.terminate()

    # ============================================================
    # Excel挿入
    # ============================================================
    print("\n■ Excelへスクリーンショット挿入中...")
    insert_screenshots()
    print("[OK] Excel更新完了")


def insert_screenshots():
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage

    # --- 006: 名寄せテスト証跡.xlsx ---
    wb6 = openpyxl.load_workbook(str(EXCEL006))
    _insert_to_sheet(wb6, "UT-NAMEYOSE-001画面証跡",
                     str(SS006 / "nameyose_001_hanyu_groupA.png"))
    _insert_to_sheet(wb6, "UT-NAMEYOSE-002画面証跡",
                     str(SS006 / "nameyose_002_hanyu_all_groups.png"))
    _insert_to_sheet(wb6, "UT-NAMEYOSE-003画面証跡",
                     str(SS006 / "nameyose_003_ohtani_groupA.png"))
    wb6.save(str(EXCEL006))
    print(f"  保存: {EXCEL006}")

    # --- 007: 加入状況テスト証跡.xlsx ---
    wb7 = openpyxl.load_workbook(str(EXCEL007))
    _insert_to_sheet(wb7, "UT-CUSTOMER-001画面証跡",
                     str(SS007 / "customer_001_hanyu_status.png"))
    _insert_to_sheet(wb7, "UT-CUSTOMER-002画面証跡",
                     str(SS007 / "customer_002_hanyu_detail.png"))
    _insert_to_sheet(wb7, "UT-CUSTOMER-003画面証跡",
                     str(SS007 / "customer_003_maturity_transition.png"))
    wb7.save(str(EXCEL007))
    print(f"  保存: {EXCEL007}")


def _insert_to_sheet(wb, sheet_name, img_path):
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    if sheet_name not in wb.sheetnames:
        print(f"  シートが見つかりません: {sheet_name}")
        return
    if not os.path.exists(img_path):
        print(f"  画像が見つかりません: {img_path}")
        return

    ws = wb[sheet_name]

    # 既存の画像を削除
    ws._images = []

    # A8セルのプレースホルダーテキストをクリア
    if ws["A8"].value and "スクリーンショット" in str(ws["A8"].value):
        ws["A8"].value = None

    # 画像をA8セルに挿入（シートの高さに合わせてリサイズ）
    img_pil = PILImage.open(img_path)
    orig_w, orig_h = img_pil.size
    target_w = 900  # ピクセル幅
    ratio = target_w / orig_w
    target_h = int(orig_h * ratio)

    img_pil_resized = img_pil.resize((target_w, target_h), PILImage.LANCZOS)
    tmp_path = img_path.replace(".png", "_thumb.png")
    img_pil_resized.save(tmp_path)

    xl_img = XLImage(tmp_path)
    xl_img.anchor = "A8"
    ws.add_image(xl_img)

    # 行高さ調整（ピクセル→ポイント: 1pt ≈ 0.75px）
    row_height_pt = target_h * 0.75
    ws.row_dimensions[8].height = row_height_pt

    print(f"  挿入完了: {sheet_name} ← {os.path.basename(img_path)}")


if __name__ == "__main__":
    run()
