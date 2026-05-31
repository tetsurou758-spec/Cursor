"""スクリーンショットをExcelに挿入するスクリプト（単独実行用）"""
import sys, os, pathlib
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

BASE_DIR = pathlib.Path(r"C:\Users\yoshi\OneDrive\ドキュメント\Cursor")
SS006    = BASE_DIR / "test" / "006_UT-名寄せバッチ処理テスト" / "screenshots"
SS007    = BASE_DIR / "test" / "007_UT-顧客管理保険加入状況テスト" / "screenshots"
EXCEL006 = BASE_DIR / "test" / "006_UT-名寄せバッチ処理テスト" / "名寄せテスト証跡.xlsx"
EXCEL007 = BASE_DIR / "test" / "007_UT-顧客管理保険加入状況テスト" / "加入状況テスト証跡.xlsx"


def insert_to_sheet(wb, sheet_name, img_path, target_w=900):
    if sheet_name not in wb.sheetnames:
        print(f"  シートが見つかりません: {sheet_name}")
        return
    if not os.path.exists(img_path):
        print(f"  画像が見つかりません: {img_path}")
        return

    ws = wb[sheet_name]
    ws._images = []

    # A8プレースホルダーをクリア
    if ws["A8"].value and "スクリーンショット" in str(ws["A8"].value):
        ws["A8"].value = None

    img_pil = PILImage.open(img_path)
    orig_w, orig_h = img_pil.size
    ratio = target_w / orig_w
    target_h = int(orig_h * ratio)

    img_pil_r = img_pil.resize((target_w, target_h), PILImage.LANCZOS)
    tmp_path = str(img_path).replace(".png", "_thumb.png")
    img_pil_r.save(tmp_path)

    xl_img = XLImage(tmp_path)
    xl_img.anchor = "A8"
    ws.add_image(xl_img)
    ws.row_dimensions[8].height = target_h * 0.75

    print(f"  挿入: {sheet_name} <- {os.path.basename(img_path)}")


# --- 006 ---
print("=== 006: 名寄せテスト証跡.xlsx ===")
wb6 = openpyxl.load_workbook(str(EXCEL006))
insert_to_sheet(wb6, "UT-NAMEYOSE-001画面証跡", str(SS006 / "nameyose_001_hanyu_groupA.png"))
insert_to_sheet(wb6, "UT-NAMEYOSE-002画面証跡", str(SS006 / "nameyose_002_hanyu_all_groups.png"))
insert_to_sheet(wb6, "UT-NAMEYOSE-003画面証跡", str(SS006 / "nameyose_003_ohtani_groupA.png"))
wb6.save(str(EXCEL006))
print(f"保存完了: {EXCEL006}")

# --- 007 ---
print("\n=== 007: 加入状況テスト証跡.xlsx ===")
wb7 = openpyxl.load_workbook(str(EXCEL007))
insert_to_sheet(wb7, "UT-CUSTOMER-001画面証跡", str(SS007 / "customer_001_hanyu_status.png"))
insert_to_sheet(wb7, "UT-CUSTOMER-002画面証跡", str(SS007 / "customer_002_hanyu_detail.png"))
insert_to_sheet(wb7, "UT-CUSTOMER-003画面証跡", str(SS007 / "customer_003_maturity_transition.png"))
wb7.save(str(EXCEL007))
print(f"保存完了: {EXCEL007}")

print("\n全完了")
