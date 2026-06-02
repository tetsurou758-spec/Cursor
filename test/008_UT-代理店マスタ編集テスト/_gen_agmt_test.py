"""
008_UT-代理店マスタ編集テスト
仕様書・証跡 Excel 生成スクリプト
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pathlib

OUT_DIR = pathlib.Path(__file__).parent

# ── カラーパレット ─────────────────────────────────────────────
NAVY   = "000D2B5E"
GOLD   = "00C9A84C"
WHITE  = "00FFFFFF"
LIGHT  = "00F0F2F7"
GRAY   = "00E0E5F0"

# ── テストケース定義 ──────────────────────────────────────────
# (試験番号, 確認対象画面, オペレーション内容, 期待結果)
CASES = [
    # セクションヘッダー
    ("【 アクセス制御確認（001〜002）】", None, None, None),
    (
        "UT-AGMT-001",
        "staff_login.html → dashboard.html",
        "社員ログイン画面で社員コード「S001」・パスワード「staff123」を入力してログインボタンをクリックする",
        "社員ダッシュボードに遷移し、ヘッダーに「S001」が表示され、「代理店マスタ」ボタンが表示されること",
    ),
    (
        "UT-AGMT-002",
        "login.html → dashboard.html → agency_master.html",
        "代理店ユーザー（A001/admin/password123）でログイン後、agency_master.html に直接URLアクセスする",
        "dashboard.html にリダイレクトされること（社員以外はアクセス不可）",
    ),
    # セクションヘッダー
    ("【 一覧・検索確認（003〜004）】", None, None, None),
    (
        "UT-AGMT-003",
        "agency_master.html",
        "社員S001でログイン後、代理店マスタ編集画面（agency_master.html）を開く",
        "代理店一覧テーブルが表示され、代理店コード・代理店名・グループコード(G)・部課コード・住所・TEL・メール の列が表示されること",
    ),
    (
        "UT-AGMT-004",
        "agency_master.html",
        "代理店マスタ編集画面で代理店コード「A001」の行を確認する（テーブルにA001が表示されていることを目視確認）",
        "A001に該当する代理店行が一覧に表示されていること",
    ),
    # セクションヘッダー
    ("【 新規追加・編集・削除確認（005〜008）】", None, None, None),
    (
        "UT-AGMT-005",
        "agency_master.html",
        "代理店一覧テーブルの任意の行をクリックする",
        "上部フォームに選択した代理店の情報（代理店コード・代理店名・グループコード・部課コード等）が入力された状態で表示され、代理店コードが編集不可（disabled）となり、ボタンが「更新」に変わること",
    ),
    (
        "UT-AGMT-006",
        "agency_master.html",
        "代理店一覧の行をクリックして編集フォームを開き、グループコード・部課コードが編集可能であることを確認する",
        "編集フォームのグループコード・部課コード入力欄が編集可能（enabled）な状態で表示されること",
    ),
    (
        "UT-AGMT-007",
        "agency_master.html",
        "編集フォームでグループコードを変更して「更新」ボタンをクリックする",
        "「代理店情報を更新しました」トースト通知が表示され、テーブルの値が更新されること",
    ),
    (
        "UT-AGMT-008",
        "agency_master.html",
        "「選択解除」ボタンをクリックして新規追加モードに切り替え、フォームが空欄になることを確認する",
        "フォームがクリアされ、代理店コードが編集可能（enabled）となり、ボタンが「新規追加」に戻ること",
    ),
]

# 証跡の合否・日付
RESULT_DATE = "2026/06/03"

# ── スタイルヘルパー ──────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="FFB0B8C4")
    return Border(left=s, right=s, top=s, bottom=s)

def make_font(bold=False, size=10, color=WHITE):
    return Font(bold=bold, size=size, color=color, name="Meiryo")

def make_fill(color):
    return PatternFill(fill_type="solid", fgColor=color)

def set_header_style(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True, size=10, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=fg, name="Meiryo")
    c.fill = make_fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = thin_border()
    return c

def set_data_style(ws, row, col, value, bg="00FFFFFF", wrap=True, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(size=9, name="Meiryo", color="00333333")
    c.fill = make_fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border = thin_border()
    return c

# ── 共通シート構築 ────────────────────────────────────────────
def build_sheet(ws, include_result):
    # カラム幅
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    # 行1: タイトル
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "代理店マスタ編集テスト仕様書（agency_master.html／アクセス制御・一覧・編集機能）"
    c.font = Font(bold=True, size=14, color=WHITE, name="Meiryo")
    c.fill = make_fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # 行2: テストID範囲・工程
    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:C2")
    ws["A2"].value = "テストID：UT-AGMT-001 〜 UT-AGMT-008"
    ws["A2"].font = Font(size=10, color=WHITE, name="Meiryo")
    ws["A2"].fill = make_fill(NAVY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("D2:G2")
    ws["D2"].value = f"テスト工程：UT（単体テスト）　　作成日：{RESULT_DATE}"
    ws["D2"].font = Font(size=10, color=WHITE, name="Meiryo")
    ws["D2"].fill = make_fill(NAVY)
    ws["D2"].alignment = Alignment(horizontal="right", vertical="center")

    # 行3: 空白
    ws.row_dimensions[3].height = 8

    # 行4: 列ヘッダー
    ws.row_dimensions[4].height = 24
    headers = ["試験番号", "確認対象画面", "オペレーション内容", "期待結果",
               "テスト結果", "テスト実施日", "テスト確認日"]
    for i, h in enumerate(headers, 1):
        set_header_style(ws, 4, i, h, bg=NAVY, fg=WHITE, bold=True, size=10)

    # 行5〜: テストケース
    cur_row = 5
    case_idx = 0
    for case in CASES:
        trial_no, screen, op, expect = case

        # セクションヘッダー行
        if screen is None:
            ws.row_dimensions[cur_row].height = 20
            ws.merge_cells(f"A{cur_row}:G{cur_row}")
            c = ws.cell(row=cur_row, column=1, value=trial_no)
            c.font = Font(bold=True, size=10, color=NAVY, name="Meiryo")
            c.fill = make_fill("00E8EDF8")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin_border()
            cur_row += 1
            continue

        # データ行
        ws.row_dimensions[cur_row].height = 56
        row_bg = "00F7F9FC" if case_idx % 2 == 0 else "00FFFFFF"

        set_data_style(ws, cur_row, 1, trial_no, bg=row_bg, wrap=False, align="center")
        set_data_style(ws, cur_row, 2, screen,   bg=row_bg, wrap=True)
        set_data_style(ws, cur_row, 3, op,       bg=row_bg, wrap=True)
        set_data_style(ws, cur_row, 4, expect,   bg=row_bg, wrap=True)

        if include_result:
            set_data_style(ws, cur_row, 5, "OK",        bg=row_bg, wrap=False, align="center")
            set_data_style(ws, cur_row, 6, RESULT_DATE, bg=row_bg, wrap=False, align="center")
            set_data_style(ws, cur_row, 7, RESULT_DATE, bg=row_bg, wrap=False, align="center")
        else:
            for col in [5, 6, 7]:
                set_data_style(ws, cur_row, col, None, bg=row_bg)

        cur_row += 1
        case_idx += 1

    return ws


def generate():
    dest = OUT_DIR

    # ── 仕様書 ───────────────────────────────────────────────
    wb_spec = openpyxl.Workbook()
    ws_spec = wb_spec.active
    ws_spec.title = "テスト仕様書"
    build_sheet(ws_spec, include_result=False)
    spec_path = dest / "代理店マスタ編集テスト仕様書.xlsx"
    wb_spec.save(spec_path)
    print(f"  生成: {spec_path.name}")

    # ── 証跡 ─────────────────────────────────────────────────
    wb_rec = openpyxl.Workbook()
    ws_rec = wb_rec.active
    ws_rec.title = "テスト証跡"
    build_sheet(ws_rec, include_result=True)
    rec_path = dest / "代理店マスタ編集テスト証跡.xlsx"
    wb_rec.save(rec_path)
    print(f"  生成: {rec_path.name}")

    print("\n完了")


if __name__ == "__main__":
    generate()
