"""
代理店マスタ編集テスト証跡.xlsx に UT-AGMT-001〜008 の個別シートを追加し、
スクリーンショットを貼り付けるスクリプト
（004_UT-権限管理テスト/_insert_screenshots.py と同一フォーマット）
"""
import io, pathlib
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_PATH = pathlib.Path(__file__).parent / "代理店マスタ編集テスト証跡.xlsx"
SS_DIR     = pathlib.Path(__file__).parent / "screenshots"

# ── 画像サイズ ────────────────────────────────────────────────
IMG_W = 1400
IMG_H = 860

# ── テストケース定義 ──────────────────────────────────────────
# (シート名, 行1タイトル, 行2確認内容, 使用スクリーンショット)
CASES = [
    (
        "UT-AGMT-001",
        "UT-AGMT-001　社員ログイン画面表示確認",
        "確認内容：社員ログイン画面（staff_login.html）でS001/staff123でログインし、社員ダッシュボードへ遷移することを確認",
        "01_staff_login.png",
    ),
    (
        "UT-AGMT-002",
        "UT-AGMT-002　代理店ユーザー アクセス制御確認",
        "確認内容：代理店ユーザー（A001/admin）でログイン後、agency_master.html に直接アクセスした際にdashboard.htmlにリダイレクトされることを確認",
        "09_agency_user_no_access.png",
    ),
    (
        "UT-AGMT-003",
        "UT-AGMT-003　代理店マスタ一覧表示確認",
        "確認内容：社員S001でagency_master.htmlを開き、代理店コード・代理店名・グループコード・部課コード等の列を持つ一覧テーブルが表示されることを確認",
        "03_agency_master_list.png",
    ),
    (
        "UT-AGMT-004",
        "UT-AGMT-004　代理店一覧 A001行表示確認",
        "確認内容：代理店マスタ一覧にA001が表示されていることを確認",
        "04_agency_search.png",
    ),
    (
        "UT-AGMT-005",
        "UT-AGMT-005　編集フォーム表示確認（行選択）",
        "確認内容：一覧の行をクリックすると上部フォームに代理店情報が入力され、代理店コードがdisabledになり「更新」ボタンに切り替わることを確認",
        "05_agency_edit_modal.png",
    ),
    (
        "UT-AGMT-006",
        "UT-AGMT-006　グループコード・部課コード 編集可能確認",
        "確認内容：編集フォームにてグループコード・部課コードが編集可能（enabled）な状態であることを確認",
        "05_agency_edit_modal.png",
    ),
    (
        "UT-AGMT-007",
        "UT-AGMT-007　編集内容保存 トースト表示確認",
        "確認内容：「更新」ボタンをクリックすると「代理店情報を更新しました」トーストが表示されることを確認",
        "06_agency_edit_save.png",
    ),
    (
        "UT-AGMT-008",
        "UT-AGMT-008　選択解除 新規追加モード切替確認",
        "確認内容：「選択解除」ボタンをクリックするとフォームがクリアされ、「新規追加」ボタンに戻ることを確認",
        "07_agency_add_modal.png",
    ),
]

# ── スタイル定数 ──────────────────────────────────────────────
NAVY     = "000D2B5E"
GOLD_BG  = "00FFF8E6"
WHITE    = "00FFFFFF"
GRAY_TXT = "005A6480"
NAVY_TXT = "000D2B5E"
EXEC_INFO = "テスト実施日：2026/06/03　　環境：ローカルサーバー（http://localhost:8000）　　ブラウザ：Chromium (Playwright headless)"


def resize_to_buf(img_path: pathlib.Path, w: int, h: int) -> io.BytesIO:
    with PILImage.open(img_path) as im:
        im = im.convert("RGB")
        im = im.resize((w, h), PILImage.LANCZOS)
        buf = io.BytesIO()
        im.save(buf, format="PNG", optimize=True)
        buf.seek(0)
    return buf


def build_case_sheet(wb, sheet_name, row1_title, row2_content, img_path):
    ws = wb.create_sheet(sheet_name)

    # カラム幅
    ws.column_dimensions["A"].width = 18.0

    # ── 行1：タイトル（navy背景・白太字）────────────────────
    ws.row_dimensions[1].height = 28.0
    ws.merge_cells("A1:H1")
    c1 = ws["A1"]
    c1.value = row1_title
    c1.font  = Font(bold=True, size=12, color=WHITE, name="Meiryo")
    c1.fill  = PatternFill(fill_type="solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center")

    # ── 行2：確認内容（薄黄背景・紺文字）────────────────────
    ws.row_dimensions[2].height = 20.0
    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = row2_content
    c2.font  = Font(bold=False, size=10, color=NAVY_TXT, name="Meiryo")
    c2.fill  = PatternFill(fill_type="solid", fgColor=GOLD_BG)
    c2.alignment = Alignment(horizontal="left", vertical="center")

    # ── 行3：実施情報（グレー文字・背景なし）────────────────
    ws.row_dimensions[3].height = 16.0
    ws.merge_cells("A3:H3")
    c3 = ws["A3"]
    c3.value = EXEC_INFO
    c3.font  = Font(bold=False, size=9, color=GRAY_TXT, name="Meiryo")
    c3.alignment = Alignment(horizontal="left", vertical="center")

    # ── 行5〜：スクリーンショット ────────────────────────────
    if img_path.exists():
        buf    = resize_to_buf(img_path, IMG_W, IMG_H)
        xl_img = XLImage(buf)
        xl_img.width  = IMG_W
        xl_img.height = IMG_H
        ws.add_image(xl_img, "A5")
    else:
        ws["A5"].value = f"（スクリーンショット未取得：{img_path.name}）"

    return ws


def main():
    wb = load_workbook(EXCEL_PATH)

    # 既存の個別ケースシートを削除（再実行対応）
    for case_id, *_ in CASES:
        if case_id in wb.sheetnames:
            del wb[case_id]
            print(f"  既存シート「{case_id}」を削除")

    # 各テストケースシートを追加
    for sheet_name, row1, row2, img_name in CASES:
        img_path = SS_DIR / img_name
        build_case_sheet(wb, sheet_name, row1, row2, img_path)
        status = "OK" if img_path.exists() else "画像なし"
        print(f"  [{sheet_name}] ← {img_name} ({status})")

    wb.save(EXCEL_PATH)
    print(f"\n保存完了: {EXCEL_PATH}")
    print(f"シート構成: {wb.sheetnames}")


if __name__ == "__main__":
    main()
