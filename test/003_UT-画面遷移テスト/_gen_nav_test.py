"""
画面遷移テスト仕様書・証跡 生成スクリプト
  - 画面遷移テスト仕様書.xlsx（Sheet1：テスト仕様書 18件）
  - 画面遷移テスト証跡.xlsx（Sheet1：DBダンプ_contracts / Sheet2：DBダンプ_users / Sheet3〜：画面キャプチャ）
実行後は不要であれば削除してよい
"""
import os, time, datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from playwright.sync_api import sync_playwright
import sqlite3

# ── パス定数 ──────────────────────────────────────────────────────
BASE_DIR  = Path(__file__).parent.parent.parent
DB_PATH   = str(BASE_DIR / "db" / "users.sqlite")
TEST_DIR  = Path(__file__).parent
SS_DIR    = TEST_DIR / "screenshots"
SPEC_PATH = TEST_DIR / "画面遷移テスト仕様書.xlsx"
EVID_PATH = TEST_DIR / "画面遷移テスト証跡.xlsx"
LOGIN_URL = (BASE_DIR / "frontend" / "login.html").as_uri()
DASH_URL  = (BASE_DIR / "frontend" / "dashboard.html").as_uri()
MAT_URL   = (BASE_DIR / "frontend" / "maturity.html").as_uri()
TODAY     = datetime.date.today().strftime("%Y/%m/%d")

SS_DIR.mkdir(exist_ok=True)

# ── スタイル ──────────────────────────────────────────────────────
NAVY  = "0D2B5E"
GOLD  = "C9A84C"
WHITE = "FFFFFF"
LIGHT = "EEF2FA"
AMBER = "FFF8E6"
BLUE2 = "2D4F8E"

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(c):   return PatternFill("solid", fgColor=c)
def center():  return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():    return Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── テストケース定義（18件）─────────────────────────────────────
# (明細番号, 確認対象画面, オペレーション内容, 期待結果)
SPEC_CASES = [
    # ── 画面遷移確認 ──
    ("UT-FLOW-001", "login.html",
     "ブラウザで login.html を開く",
     "ログイン画面が表示されること（AX損害保険株式会社のログインフォームが見えること）"),
    ("UT-FLOW-002", "login.html → dashboard.html",
     "代理店コード「A001」・ログインID「admin」・パスワード「password123」を入力してログインボタンをクリックする",
     "dashboard.html へ遷移し、ヘッダーに「A001」「admin」が表示されること"),
    ("UT-FLOW-003", "dashboard.html → maturity.html",
     "ダッシュボードの「満期管理システム」ボタンをクリックする",
     "maturity.html へ遷移し、満期管理一覧が表示されること"),
    ("UT-FLOW-004", "maturity.html → dashboard.html",
     "満期管理画面の「← ダッシュボード」ボタンをクリックする",
     "dashboard.html に戻り、ダッシュボードが表示されること"),
    ("UT-FLOW-005", "login.html → dashboard.html",
     "代理店コード「B002」・ログインID「agent1」・パスワード「pass456」を入力してログインボタンをクリックする",
     "dashboard.html へ遷移し、ヘッダーに「B002」「agent1」が表示されること"),
    ("UT-FLOW-006", "dashboard.html → maturity.html",
     "B002でログイン後、ダッシュボードの「満期管理システム」ボタンをクリックする",
     "maturity.html へ遷移し、B002の満期管理一覧が表示されること"),
    ("UT-FLOW-007", "login.html → dashboard.html",
     "代理店コード「C003」・ログインID「user1」・パスワード「pass789」を入力してログインボタンをクリックする",
     "dashboard.html へ遷移し、ヘッダーに「C003」「user1」が表示されること"),
    ("UT-FLOW-008", "dashboard.html → maturity.html",
     "C003でログイン後、ダッシュボードの「満期管理システム」ボタンをクリックする",
     "maturity.html へ遷移し、C003の満期管理一覧が表示されること"),
    # ── ダッシュボード表示データ整合性確認 ──
    ("UT-FLOW-009", "dashboard.html / GET /api/dashboard",
     "A001でログインしてダッシュボードを表示し、今月更改総件数を確認する",
     "A001の今月更改総件数が DB の COUNTIFS 結果「247件」と一致すること"),
    ("UT-FLOW-010", "dashboard.html / GET /api/dashboard",
     "A001でログインしてダッシュボードを表示し、今月更改済件数を確認する",
     "A001の今月更改済件数が DB の COUNTIFS 結果「28件」と一致すること"),
    ("UT-FLOW-011", "dashboard.html / GET /api/dashboard",
     "A001でログインしてダッシュボードを表示し、翌月更改総件数を確認する",
     "A001の翌月更改総件数が DB の COUNTIFS 結果「325件」と一致すること"),
    ("UT-FLOW-012", "dashboard.html / GET /api/dashboard",
     "B002でログインしてダッシュボードを表示し、今月更改総件数を確認する",
     "B002の今月更改総件数が DB の COUNTIFS 結果「197件」と一致すること"),
    ("UT-FLOW-013", "dashboard.html / GET /api/dashboard",
     "C003でログインしてダッシュボードを表示し、今月更改総件数を確認する",
     "C003の今月更改総件数が DB の COUNTIFS 結果「157件」と一致すること"),
    # ── 満期管理画面データ整合性確認 ──
    ("UT-FLOW-014", "maturity.html / GET /api/maturity",
     "A001でログインして満期管理画面を表示し、一覧件数を確認する",
     "A001の満期管理一覧件数が DB の COUNTIFS 結果「5件」と一致すること"),
    ("UT-FLOW-015", "maturity.html / GET /api/maturity",
     "B002でログインして満期管理画面を表示し、一覧件数を確認する",
     "B002の満期管理一覧件数が DB の COUNTIFS 結果「5件」と一致すること"),
    ("UT-FLOW-016", "maturity.html / GET /api/maturity",
     "C003でログインして満期管理画面を表示し、一覧件数を確認する",
     "C003の満期管理一覧件数が DB の COUNTIFS 結果「5件」と一致すること"),
    ("UT-FLOW-017", "maturity.html / GET /api/maturity",
     "A001でログインして満期管理画面でフォローコール「未実施」を選択し検索する",
     "A001のフォローコール未実施件数が DB の COUNTIFS 結果「2件」と一致すること"),
    ("UT-FLOW-018", "maturity.html / GET /api/maturity",
     "A001でログインして満期管理画面で更改STS「落ち」を選択し検索する",
     "A001の更改STS「落ち」件数が DB の COUNTIFS 結果「1件」と一致すること"),
]

COLUMNS = ["明細番号", "確認対象画面（遷移元→遷移先）", "オペレーション内容", "期待結果", "テスト結果", "テスト実施日", "テスト確認日"]

GROUPS = [
    ("画面遷移確認（001〜008）",                          8),
    ("ダッシュボード表示データ整合性確認（009〜013）",    5),
    ("満期管理画面データ整合性確認（014〜018）",          5),
]


# ════════════════════════════════════════════════════════
# 1. テスト仕様書生成
# ════════════════════════════════════════════════════════
def gen_spec():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "テスト仕様書"
    LAST = get_column_letter(len(COLUMNS))

    # 行1: タイトル
    ws.merge_cells(f"A1:{LAST}1")
    c = ws["A1"]
    c.value     = "ログイン→ダッシュボード→満期管理　画面遷移テスト仕様書"
    c.font      = Font(name="游ゴシック", bold=True, size=14, color=WHITE)
    c.fill      = fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 30

    # 行2: メタ情報
    ws.merge_cells("A2:C2")
    m1 = ws["A2"]
    m1.value     = f"テストID：UT-FLOW-001 ～ UT-FLOW-{len(SPEC_CASES):03d}　　テスト確認観点：画面遷移・データ整合性全条件網羅テスト"
    m1.font      = Font(name="游ゴシック", bold=True, size=10, color=NAVY)
    m1.fill      = fill(AMBER)
    m1.alignment = left()
    ws.merge_cells(f"D2:{LAST}2")
    m2 = ws["D2"]
    m2.value     = f"テスト工程：UT（単体テスト）　　作成日：{TODAY}"
    m2.font      = Font(name="游ゴシック", bold=True, size=10, color=NAVY)
    m2.fill      = fill(AMBER)
    m2.alignment = left()
    ws.row_dimensions[2].height = 22

    # 行3: ゴールドライン
    for col in range(1, len(COLUMNS) + 1):
        ws.cell(row=3, column=col).fill = fill(GOLD)
    ws.row_dimensions[3].height = 4

    # 行4: ヘッダー
    for ci, name in enumerate(COLUMNS, 1):
        c = ws.cell(row=4, column=ci, value=name)
        c.font      = Font(name="游ゴシック", bold=True, size=10, color=WHITE)
        c.fill      = fill(NAVY)
        c.alignment = center()
        c.border    = thin_border()
    ws.row_dimensions[4].height = 24

    # 行5〜: データ行
    for ri, case in enumerate(SPEC_CASES, 5):
        bg = WHITE if ri % 2 == 1 else LIGHT
        row_data = list(case) + ["", "", ""]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border    = thin_border()
            c.fill      = fill(bg)
            c.font      = Font(name="游ゴシック", size=10, color="1A1A2E")
            c.alignment = center() if ci in (1, 5, 6, 7) else left()
        ws.row_dimensions[ri].height = 38

    # グループ区切り行を挿入（逆順で行番号ズレを防ぐ）
    group_insert = []
    cursor = 5
    for label, count in GROUPS:
        group_insert.append((cursor, f"■ {label}"))
        cursor += count
    for gr, label in reversed(group_insert):
        ws.insert_rows(gr)
        ws.merge_cells(f"A{gr}:{LAST}{gr}")
        gc = ws[f"A{gr}"]
        gc.value     = label
        gc.font      = Font(name="游ゴシック", bold=True, size=9, color=WHITE)
        gc.fill      = fill(BLUE2)
        gc.alignment = left()
        ws.row_dimensions[gr].height = 18

    # 列幅
    for i, w in enumerate([18, 32, 54, 52, 12, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows = "1:4"

    wb.save(str(SPEC_PATH))
    print(f"仕様書生成完了: {SPEC_PATH}")


# ════════════════════════════════════════════════════════
# 2. Playwright でスクリーンショット取得
# ════════════════════════════════════════════════════════
def take_screenshots() -> dict:
    """A001/B002/C003 の3フロー分スクリーンショットを取得する"""
    paths = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx     = browser.new_context(viewport={"width": 1400, "height": 860})
        page    = ctx.new_page()

        # ═══ A001 フロー ════════════════════════════════
        # UT-FLOW-001: ログイン画面表示
        print("  UT-FLOW-001: ログイン画面")
        page.goto(LOGIN_URL)
        page.wait_for_load_state("domcontentloaded")
        page.evaluate("localStorage.clear()")
        page.reload()
        page.wait_for_load_state("domcontentloaded")
        ss = str(SS_DIR / "UT-FLOW-001.png")
        page.screenshot(path=ss)
        paths["UT-FLOW-001"] = ss

        # UT-FLOW-002: A001 ログイン → ダッシュボード
        print("  UT-FLOW-002: A001 ログイン → ダッシュボード")
        page.fill("#agency-code", "A001")
        page.fill("#login-id",    "admin")
        page.fill("#password",    "password123")
        page.click("button[type='submit']")
        page.wait_for_url("**/dashboard.html", timeout=10000)
        time.sleep(2.5)
        ss = str(SS_DIR / "UT-FLOW-002.png")
        page.screenshot(path=ss)
        paths["UT-FLOW-002"] = ss

        # UT-FLOW-003: 満期管理ボタン → 満期管理
        print("  UT-FLOW-003: 満期管理ボタン → maturity.html")
        page.click("a[href='maturity.html']")
        page.wait_for_url("**/maturity.html", timeout=10000)
        time.sleep(2.5)
        ss = str(SS_DIR / "UT-FLOW-003.png")
        page.screenshot(path=ss)
        paths["UT-FLOW-003"] = ss

        # UT-FLOW-004: ← ダッシュボード → ダッシュボード戻り
        print("  UT-FLOW-004: ← ダッシュボード → dashboard.html")
        page.click("a.btn-back")
        page.wait_for_url("**/dashboard.html", timeout=10000)
        time.sleep(2.0)
        ss = str(SS_DIR / "UT-FLOW-004.png")
        page.screenshot(path=ss)
        paths["UT-FLOW-004"] = ss

        # ═══ B002 フロー ════════════════════════════════
        # UT-FLOW-005: B002 ログイン → ダッシュボード
        print("  UT-FLOW-005: B002 ログイン → ダッシュボード")
        page.click("#btn-logout")
        page.wait_for_url("**/login.html", timeout=8000)
        page.wait_for_load_state("domcontentloaded")
        page.fill("#agency-code", "B002")
        page.fill("#login-id",    "agent1")
        page.fill("#password",    "pass456")
        page.click("button[type='submit']")
        page.wait_for_url("**/dashboard.html", timeout=10000)
        time.sleep(2.5)
        ss = str(SS_DIR / "UT-FLOW-005.png")
        page.screenshot(path=ss)
        paths["UT-FLOW-005"] = ss

        # UT-FLOW-006: B002 満期管理
        print("  UT-FLOW-006: B002 満期管理")
        page.click("a[href='maturity.html']")
        page.wait_for_url("**/maturity.html", timeout=10000)
        time.sleep(2.5)
        ss = str(SS_DIR / "UT-FLOW-006.png")
        page.screenshot(path=ss)
        paths["UT-FLOW-006"] = ss

        # ═══ C003 フロー ════════════════════════════════
        # UT-FLOW-007: C003 ログイン → ダッシュボード
        print("  UT-FLOW-007: C003 ログイン → ダッシュボード")
        page.click("a.btn-back")
        page.wait_for_url("**/dashboard.html", timeout=8000)
        page.click("#btn-logout")
        page.wait_for_url("**/login.html", timeout=8000)
        page.wait_for_load_state("domcontentloaded")
        page.fill("#agency-code", "C003")
        page.fill("#login-id",    "user1")
        page.fill("#password",    "pass789")
        page.click("button[type='submit']")
        page.wait_for_url("**/dashboard.html", timeout=10000)
        time.sleep(2.5)
        ss = str(SS_DIR / "UT-FLOW-007.png")
        page.screenshot(path=ss)
        paths["UT-FLOW-007"] = ss

        # UT-FLOW-008: C003 満期管理
        print("  UT-FLOW-008: C003 満期管理")
        page.click("a[href='maturity.html']")
        page.wait_for_url("**/maturity.html", timeout=10000)
        time.sleep(2.5)
        ss = str(SS_DIR / "UT-FLOW-008.png")
        page.screenshot(path=ss)
        paths["UT-FLOW-008"] = ss

        browser.close()

    return paths


# ════════════════════════════════════════════════════════
# 3. テスト証跡生成
# ════════════════════════════════════════════════════════
SS_META = {
    "UT-FLOW-001": ("UT-FLOW-001　ログイン画面表示",
                    "login.html を開いてログインフォームが表示されることを確認"),
    "UT-FLOW-002": ("UT-FLOW-002　A001 ログイン成功 → ダッシュボード遷移",
                    "A001/admin でログイン後、dashboard.html へ遷移しヘッダーに代理店コード・ログインIDが表示されることを確認"),
    "UT-FLOW-003": ("UT-FLOW-003　A001 ダッシュボード → 満期管理画面遷移",
                    "「満期管理システム」ボタン押下で maturity.html へ遷移し一覧が表示されることを確認"),
    "UT-FLOW-004": ("UT-FLOW-004　A001 満期管理 → ダッシュボード（戻り）",
                    "「← ダッシュボード」ボタン押下で dashboard.html へ戻ったことを確認"),
    "UT-FLOW-005": ("UT-FLOW-005　B002 ログイン成功 → ダッシュボード遷移",
                    "B002/agent1 でログイン後、dashboard.html へ遷移しヘッダーに B002 が表示されることを確認"),
    "UT-FLOW-006": ("UT-FLOW-006　B002 ダッシュボード → 満期管理画面遷移",
                    "B002 で maturity.html へ遷移し B002 の満期一覧が表示されることを確認"),
    "UT-FLOW-007": ("UT-FLOW-007　C003 ログイン成功 → ダッシュボード遷移",
                    "C003/user1 でログイン後、dashboard.html へ遷移しヘッダーに C003 が表示されることを確認"),
    "UT-FLOW-008": ("UT-FLOW-008　C003 ダッシュボード → 満期管理画面遷移",
                    "C003 で maturity.html へ遷移し C003 の満期一覧が表示されることを確認"),
}


def gen_evidence(ss_paths: dict):
    wb  = openpyxl.Workbook()

    # ── Sheet1: DBダンプ_contracts ──────────────────────────────
    ws1 = wb.active
    ws1.title = "DBダンプ_contracts"

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()
    cur.execute("""
        SELECT id, agency_code, contract_no, customer_name,
               renewal_month, status, expiry_date,
               followcall_status, renewal_status, created_at
        FROM contracts
        ORDER BY agency_code, expiry_date, renewal_month
    """)
    contracts = cur.fetchall()
    total_rows = len(contracts)
    # レイアウト: 行1=タイトル,2=見出し,3=マトリクスヘッダー,4-6=代理店,7=合計,8=空白
    # 行9=テストケース対応表ヘッダー,行10-18=対応表(9行),行19=空白,行20=一覧ヘッダー,行21=データヘッダー,行22〜=データ
    DATA_START = 22
    DATA_END   = DATA_START + total_rows - 1
    B = f"$B${DATA_START}:$B${DATA_END}"   # agency_code
    E = f"$E${DATA_START}:$E${DATA_END}"   # renewal_month
    F = f"$F${DATA_START}:$F${DATA_END}"   # status
    G = f"$G${DATA_START}:$G${DATA_END}"   # expiry_date
    H = f"$H${DATA_START}:$H${DATA_END}"   # followcall_status
    I = f"$I${DATA_START}:$I${DATA_END}"   # renewal_status

    # タイトル行
    MATRIX_COLS = 10
    LAST_MCOL = get_column_letter(MATRIX_COLS)
    ws1.merge_cells(f"A1:{LAST_MCOL}1")
    t = ws1["A1"]
    t.value     = "DBダンプ：contracts テーブル 集計マトリクス（COUNTIFS）"
    t.font      = Font(name="游ゴシック", bold=True, size=13, color=WHITE)
    t.fill      = fill(NAVY)
    t.alignment = center()
    ws1.row_dimensions[1].height = 28

    # 行2: 見出し
    ws1.merge_cells(f"A2:{LAST_MCOL}2")
    ws1["A2"].value     = "集計マトリクス（代理店コード × ダッシュボード集計 × 満期管理件数）"
    ws1["A2"].font      = Font(name="游ゴシック", bold=True, size=10, color=NAVY)
    ws1["A2"].fill      = fill(AMBER)
    ws1["A2"].alignment = left()
    ws1.row_dimensions[2].height = 20

    # 行3: マトリクスヘッダー
    matrix_headers = [
        "代理店コード",
        "今月総件数\n(2026-05)",
        "今月更改済\n(completed)",
        "今月未更改\n(pending)",
        "翌月総件数\n(2026-06)",
        "翌月更改済\n(completed)",
        "翌月未更改\n(pending)",
        "満期管理\n表示件数",
        "フォローコール\n未実施件数",
        "更改STS\n落ち件数",
    ]
    for ci, h in enumerate(matrix_headers, 1):
        c = ws1.cell(row=3, column=ci, value=h)
        c.font      = Font(name="游ゴシック", bold=True, size=9, color=WHITE)
        c.fill      = fill(NAVY)
        c.alignment = center()
        c.border    = thin_border()
    ws1.row_dimensions[3].height = 32

    # 行4〜6: 各代理店の COUNTIFS 数式
    agencies = ["A001", "B002", "C003"]
    for ri, ag in enumerate(agencies, 4):
        bg = WHITE if ri % 2 == 1 else LIGHT
        ws1.cell(row=ri, column=1, value=ag).font = Font(name="游ゴシック", bold=True, size=10, color=NAVY)
        ws1.cell(row=ri, column=1).fill      = fill(bg)
        ws1.cell(row=ri, column=1).alignment = center()
        ws1.cell(row=ri, column=1).border    = thin_border()
        # 今月（ダッシュボード：renewal_month ベース）
        ws1.cell(row=ri, column=2,  value=f'=COUNTIFS({B},"{ag}",{E},"2026-05")')
        ws1.cell(row=ri, column=3,  value=f'=COUNTIFS({B},"{ag}",{E},"2026-05",{F},"completed")')
        ws1.cell(row=ri, column=4,  value=f'=COUNTIFS({B},"{ag}",{E},"2026-05",{F},"pending")')
        # 翌月
        ws1.cell(row=ri, column=5,  value=f'=COUNTIFS({B},"{ag}",{E},"2026-06")')
        ws1.cell(row=ri, column=6,  value=f'=COUNTIFS({B},"{ag}",{E},"2026-06",{F},"completed")')
        ws1.cell(row=ri, column=7,  value=f'=COUNTIFS({B},"{ag}",{E},"2026-06",{F},"pending")')
        # 満期管理（expiry_date IS NOT NULL でデフォルト期間：2026-02-01〜2026-08-28）
        ws1.cell(row=ri, column=8,  value=f'=COUNTIFS({B},"{ag}",{G},">="&"2026-02-01",{G},"<="&"2026-08-28")')
        ws1.cell(row=ri, column=9,  value=f'=COUNTIFS({B},"{ag}",{H},"未実施")')
        ws1.cell(row=ri, column=10, value=f'=COUNTIFS({B},"{ag}",{I},"落ち")')
        for ci in range(2, MATRIX_COLS + 1):
            ws1.cell(row=ri, column=ci).font      = Font(name="游ゴシック", size=10)
            ws1.cell(row=ri, column=ci).fill      = fill(bg)
            ws1.cell(row=ri, column=ci).alignment = center()
            ws1.cell(row=ri, column=ci).border    = thin_border()
        ws1.row_dimensions[ri].height = 18

    # 行7: 合計
    ws1.cell(row=7, column=1, value="合計").font = Font(name="游ゴシック", bold=True, size=10, color=WHITE)
    ws1.cell(row=7, column=1).fill      = fill(NAVY)
    ws1.cell(row=7, column=1).alignment = center()
    ws1.cell(row=7, column=1).border    = thin_border()
    for ci in range(2, MATRIX_COLS + 1):
        ws1.cell(row=7, column=ci, value=f"=SUM({get_column_letter(ci)}4:{get_column_letter(ci)}6)")
        ws1.cell(row=7, column=ci).font      = Font(name="游ゴシック", bold=True, size=10, color=WHITE)
        ws1.cell(row=7, column=ci).fill      = fill(NAVY)
        ws1.cell(row=7, column=ci).alignment = center()
        ws1.cell(row=7, column=ci).border    = thin_border()
    ws1.row_dimensions[7].height = 18

    # 行8: テストケース対応表
    ws1.row_dimensions[8].height = 8
    ws1.merge_cells(f"A9:{LAST_MCOL}9")
    ws1["A9"].value     = "テストケース対応表"
    ws1["A9"].font      = Font(name="游ゴシック", bold=True, size=10, color=WHITE)
    ws1["A9"].fill      = fill(BLUE2)
    ws1["A9"].alignment = left()
    ws1.row_dimensions[9].height = 18

    tc_map = [
        ("列", "テストケース", "確認対象"),
        ("今月総件数", "UT-FLOW-009", "A001 今月更改総件数"),
        ("今月更改済", "UT-FLOW-010", "A001 今月更改済件数"),
        ("翌月総件数", "UT-FLOW-011", "A001 翌月更改総件数"),
        ("今月総件数", "UT-FLOW-012", "B002 今月更改総件数"),
        ("今月総件数", "UT-FLOW-013", "C003 今月更改総件数"),
        ("満期管理表示", "UT-FLOW-014/015/016", "各社 満期管理一覧件数"),
        ("フォローコール未実施", "UT-FLOW-017", "A001 フォローコール未実施"),
        ("更改STS落ち", "UT-FLOW-018", "A001 更改STS落ち"),
    ]
    for ri, (col, tc, desc) in enumerate(tc_map, 10):
        bg = WHITE if ri % 2 == 0 else LIGHT
        ws1.cell(row=ri, column=1, value=col).font = Font(name="游ゴシック", bold=True, size=9, color=NAVY)
        ws1.cell(row=ri, column=1).fill = fill(bg)
        ws1.cell(row=ri, column=1).alignment = center()
        ws1.cell(row=ri, column=1).border = thin_border()
        ws1.cell(row=ri, column=2, value=tc).font = Font(name="游ゴシック", size=9, color="1A1A2E")
        ws1.cell(row=ri, column=2).fill = fill(bg)
        ws1.cell(row=ri, column=2).alignment = center()
        ws1.cell(row=ri, column=2).border = thin_border()
        ws1.merge_cells(f"C{ri}:{LAST_MCOL}{ri}")
        ws1.cell(row=ri, column=3, value=desc).font = Font(name="游ゴシック", size=9, color="1A1A2E")
        ws1.cell(row=ri, column=3).fill = fill(bg)
        ws1.cell(row=ri, column=3).alignment = left()
        ws1.cell(row=ri, column=3).border = thin_border()
        ws1.row_dimensions[ri].height = 16

    # 全レコード一覧 (行19=空白, 行20=一覧ヘッダー, 行21=データヘッダー)
    ws1.row_dimensions[19].height = 8
    divider_row = 20
    ws1.merge_cells(f"A{divider_row}:{LAST_MCOL}{divider_row}")
    ws1[f"A{divider_row}"].value     = f"全レコード一覧（{total_rows}件）"
    ws1[f"A{divider_row}"].font      = Font(name="游ゴシック", bold=True, size=10, color=WHITE)
    ws1[f"A{divider_row}"].fill      = fill(BLUE2)
    ws1[f"A{divider_row}"].alignment = left()
    ws1.row_dimensions[divider_row].height = 18

    # 全レコードヘッダー（行21）
    hdr_row = 21
    data_cols = ["id", "agency_code", "contract_no", "customer_name",
                 "renewal_month", "status", "expiry_date", "followcall_status", "renewal_status", "created_at"]
    for ci, h in enumerate(data_cols, 1):
        c = ws1.cell(row=hdr_row, column=ci, value=h)
        c.font      = Font(name="游ゴシック", bold=True, size=9, color=WHITE)
        c.fill      = fill(NAVY)
        c.alignment = center()
        c.border    = thin_border()
    ws1.row_dimensions[hdr_row].height = 18

    # 全レコード
    for ri, row in enumerate(contracts, DATA_START):
        bg = WHITE if ri % 2 == 0 else LIGHT
        for ci, val in enumerate(row, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="游ゴシック", size=8)
            c.fill      = fill(bg)
            c.alignment = center() if ci in (1, 2, 5, 6) else left()
            c.border    = thin_border()
        ws1.row_dimensions[ri].height = 14

    # 列幅
    for i, w in enumerate([12, 12, 20, 18, 12, 12, 12, 14, 12, 18], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    conn.row_factory = sqlite3.Row
    cur2 = conn.cursor()

    # ── Sheet2: DBダンプ_users ───────────────────────────────────
    ws2 = wb.create_sheet("DBダンプ_users")
    cur2.execute("SELECT id, agency_code, login_id, name, created_at FROM users ORDER BY id")
    users = cur2.fetchall()
    conn.close()

    ws2.merge_cells("A1:E1")
    t2 = ws2["A1"]
    t2.value     = "DBダンプ：users テーブル（代理店コード・ログインID 証跡）"
    t2.font      = Font(name="游ゴシック", bold=True, size=13, color=WHITE)
    t2.fill      = fill(NAVY)
    t2.alignment = center()
    ws2.row_dimensions[1].height = 28

    user_cols = ["id", "agency_code", "login_id", "name", "created_at"]
    for ci, h in enumerate(user_cols, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font      = Font(name="游ゴシック", bold=True, size=10, color=WHITE)
        c.fill      = fill(NAVY)
        c.alignment = center()
        c.border    = thin_border()
    ws2.row_dimensions[2].height = 22

    for ri, row in enumerate(users, 3):
        for ci, val in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="游ゴシック", size=10)
            c.fill      = fill(WHITE if ri % 2 == 1 else LIGHT)
            c.alignment = center() if ci in (1, 2, 3) else left()
            c.border    = thin_border()
        ws2.row_dimensions[ri].height = 20

    for i, w in enumerate([6, 14, 14, 18, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet3〜: 画面遷移キャプチャ証跡 ────────────────────────
    ordered_ids = ["UT-FLOW-001", "UT-FLOW-002", "UT-FLOW-003", "UT-FLOW-004",
                   "UT-FLOW-005", "UT-FLOW-006", "UT-FLOW-007", "UT-FLOW-008"]
    for test_id in ordered_ids:
        img_path = ss_paths.get(test_id)
        title, desc = SS_META.get(test_id, (test_id, ""))

        ws = wb.create_sheet(test_id)
        ws.merge_cells("A1:H1")
        h = ws["A1"]
        h.value     = title
        h.font      = Font(name="游ゴシック", bold=True, size=12, color=WHITE)
        h.fill      = fill(NAVY)
        h.alignment = center()
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:H2")
        d = ws["A2"]
        d.value     = f"確認内容：{desc}"
        d.font      = Font(name="游ゴシック", size=10, color=NAVY)
        d.fill      = fill(AMBER)
        d.alignment = left()
        ws.row_dimensions[2].height = 20

        ws.merge_cells("A3:H3")
        ws[f"A3"].value     = f"テスト実施日：{TODAY}　　環境：ローカルファイル（file://）　　ブラウザ：Chromium (Playwright headless)"
        ws["A3"].font       = Font(name="游ゴシック", size=9, color="5A6480")
        ws["A3"].alignment  = left()
        ws.row_dimensions[3].height = 16

        if img_path and os.path.exists(img_path):
            img        = XlImage(img_path)
            img.width  = 900
            img.height = 563
            ws.add_image(img, "A5")
            ws.row_dimensions[5].height = 430
        else:
            ws["A5"].value = "（スクリーンショット未取得）"

        ws.column_dimensions["A"].width = 18

    wb.save(str(EVID_PATH))
    print(f"証跡生成完了: {EVID_PATH}")


# ════════════════════════════════════════════════════════
# メイン
# ════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 55)
    print("  画面遷移テスト 仕様書・証跡 生成")
    print("=" * 55)

    print("\n【Step 1】テスト仕様書 生成")
    gen_spec()

    print("\n【Step 2】Playwright スクリーンショット取得（バックエンド起動必須）")
    ss_paths = take_screenshots()
    print(f"  取得完了: {len(ss_paths)}件")

    print("\n【Step 3】テスト証跡 生成")
    gen_evidence(ss_paths)

    print("\n" + "=" * 55)
    print("  完了")
    print("=" * 55)
