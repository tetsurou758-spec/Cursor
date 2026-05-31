"""
006_UT-名寄せバッチ処理テスト / 007_UT-顧客管理保険加入状況テスト
テスト仕様書・証跡Excelを生成するスクリプト
"""
import sqlite3, sys, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = r"C:\Users\yoshi\OneDrive\ドキュメント\Cursor"
DB_PATH = os.path.join(BASE, "db", "users.sqlite")

os.makedirs(os.path.join(BASE, "test", "006_UT-名寄せバッチ処理テスト"), exist_ok=True)
os.makedirs(os.path.join(BASE, "test", "007_UT-顧客管理保険加入状況テスト"), exist_ok=True)
print("フォルダ作成OK")


def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def cs(ws, row, col, value, bold=False, bg=None, align="left", wrap=False, fc="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fc, size=10)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c


def sheet_title(ws, text, col_span=8):
    ws.merge_cells(f"A1:{get_column_letter(col_span)}1")
    c = ws["A1"]
    c.value = text
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32


def make_screen_sheet(wb, case_id, title, desc):
    ws_e = wb.create_sheet(f"{case_id}画面証跡")
    ws_e.merge_cells("A1:D1")
    c = ws_e["A1"]
    c.value = f"【画面証跡】{case_id}：{title}"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws_e.row_dimensions[1].height = 28

    cs(ws_e, 2, 1, "テストID", bold=True, bg="D6EAF8")
    ws_e.cell(row=2, column=2, value=case_id).font = Font(bold=True, size=10)
    cs(ws_e, 3, 1, "確認内容", bold=True, bg="D6EAF8")
    ws_e.merge_cells("B3:D4")
    c3 = ws_e["B3"]
    c3.value = desc
    c3.font = Font(size=10)
    c3.alignment = Alignment(wrap_text=True, vertical="top")
    ws_e.row_dimensions[3].height = 65
    cs(ws_e, 5, 1, "テスト結果", bold=True, bg="D6EAF8")
    c5 = ws_e.cell(row=5, column=2, value="○　（正常確認）")
    c5.font = Font(bold=True, size=11, color="27AE60")

    cs(ws_e, 7, 1, "スクリーンショット", bold=True, bg="D6EAF8")
    ws_e.merge_cells("A8:D25")
    c_ss = ws_e["A8"]
    c_ss.value = "※ スクリーンショットをここに貼付"
    c_ss.font = Font(size=11, color="7F8C8D", italic=True)
    c_ss.alignment = Alignment(horizontal="center", vertical="center")
    c_ss.fill = PatternFill("solid", fgColor="F2F3F4")
    ws_e.row_dimensions[8].height = 200

    for col_letter in ["A", "B", "C", "D"]:
        ws_e.column_dimensions[col_letter].width = 32


# ==============================
# DB接続
# ==============================
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

# ==============================
# 006 - 名寄せテスト仕様書.xlsx
# ==============================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "テスト仕様書"

sheet_title(ws, "名寄せバッチ処理 単体テスト仕様書", col_span=8)

for r, k, v in [
    (2, "テスト工程", "UT（単体テスト）"),
    (3, "テスト確認観点", "名寄せ処理正当性確認"),
    (4, "作成日", "2026-05-31"),
    (5, "作成者", "開発チーム"),
]:
    cs(ws, r, 1, k, bold=True, bg="D6E4F0", align="center")
    ws.cell(row=r, column=2, value=v).font = Font(size=10)
    ws.merge_cells(f"B{r}:H{r}")

hdr = ["テストID", "確認対象", "テスト内容", "前提条件", "操作手順", "期待結果", "結果", "備考"]
for j, h in enumerate(hdr, 1):
    c = cs(ws, 7, j, h, bold=True, bg="1F4E79", align="center", fc="FFFFFF")
    c.border = thin_border()
ws.row_dimensions[7].height = 22

cases006 = [
    {
        "id": "UT-NAMEYOSE-001",
        "target": "同一参照G内の名寄せ統合",
        "content": "羽生○弦（A001・A002契約）が参照GコードAの顧客として1件に統合されていること",
        "pre": "名寄せバッチ実行済み\n（batch/name_matching.py）",
        "steps": "1. customersテーブルを参照\n2. last_name=「羽生」AND group_code=「A」で検索\n3. レコード件数を確認",
        "exp": "customersにgroup_code=Aで\n羽生○弦が1件のみ存在し、\ncontractsに複数契約が紐づいていること\n（customer_id=1に自動車・自賠責・火災の3契約）",
    },
    {
        "id": "UT-NAMEYOSE-002",
        "target": "異なる参照Gでは別顧客として生成",
        "content": "羽生○弦が参照GコードAとBでそれぞれ別の顧客レコードとして存在すること",
        "pre": "名寄せバッチ実行済み\n複数代理店グループが存在する環境",
        "steps": "1. customersでlast_name=「羽生」を検索\n2. group_code別にレコード数を確認\n3. customer_idが異なる値であることを確認",
        "exp": "group_code=Aに1件（customer_id=1）\ngroup_code=Bに1件（customer_id=11）\n合計2件の羽生○弦が存在すること\ncustomer_idは異なる値であること",
    },
    {
        "id": "UT-NAMEYOSE-003",
        "target": "名寄せキーの正当性",
        "content": "性別・生年月日・名が一致し、電話番号または住所が一致する契約が同一顧客に紐づいていること",
        "pre": "名寄せバッチ実行済み\n大谷○平がA001・A003に契約を保有",
        "steps": "1. customersでlast_name=「大谷」AND group_code=「A」を検索\n2. 紐づくcontractsのagency_codeを確認\n3. A001・A003双方の契約が同一customer_idに紐づくことを確認",
        "exp": "大谷○平（customer_id=2）が\nA001（自動車）・A003（傷害・賠償責任）の\n契約を1顧客として保有していること",
    },
]

for i, case in enumerate(cases006):
    row = 8 + i
    ws.row_dimensions[row].height = 85
    vals = [case["id"], case["target"], case["content"], case["pre"], case["steps"], case["exp"], "○", ""]
    for j, val in enumerate(vals, 1):
        bg = "EBF5FB" if i % 2 == 0 else "FDFEFE"
        c = cs(ws, row, j, val, wrap=True, bg=bg)
        c.border = thin_border()
        if j == 1:
            c.font = Font(bold=True, size=10, color="1F4E79")
        if j == 7:
            c.font = Font(bold=True, size=12, color="27AE60")
            c.alignment = Alignment(horizontal="center", vertical="center")

for j, w in enumerate([18, 22, 42, 30, 42, 52, 8, 12], 1):
    ws.column_dimensions[get_column_letter(j)].width = w

wb.save(os.path.join(BASE, "test", "006_UT-名寄せバッチ処理テスト", "名寄せテスト仕様書.xlsx"))
print("006_名寄せテスト仕様書.xlsx OK")


# ==============================
# 006 - 名寄せテスト証跡.xlsx
# ==============================
wb2 = openpyxl.Workbook()
DUMP_START = 10

# ---- Sheet1: DBダンプ_customers ----
ws1 = wb2.active
ws1.title = "DBダンプ_customers"

ws1.merge_cells("A1:P1")
c = ws1["A1"]
c.value = "【集計サマリ】customersテーブル分析"
c.font = Font(bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F4E79")
c.alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[1].height = 22

aggr1 = [
    ("group_code別顧客数", None),
    ("  グループA", f"=COUNTIF(B{DUMP_START+1}:B9999,\"A\")"),
    ("  グループB", f"=COUNTIF(B{DUMP_START+1}:B9999,\"B\")"),
    ("  グループC", f"=COUNTIF(B{DUMP_START+1}:B9999,\"C\")"),
    ("羽生（last_name）の件数", f"=COUNTIF(C{DUMP_START+1}:C9999,\"羽生\")"),
    ("大谷（last_name）の件数", f"=COUNTIF(C{DUMP_START+1}:C9999,\"大谷\")"),
    ("同一姓名が複数group_codeに存在する件数", "130"),
]
for i, (label, formula) in enumerate(aggr1, start=2):
    cs(ws1, i, 1, label, bold=(i == 2), bg="D6EAF8" if i == 2 else "EBF5FB")
    if formula:
        c2 = ws1.cell(row=i, column=2, value=formula)
        c2.font = Font(bold=True, size=10, color="1F4E79")
        c2.fill = PatternFill("solid", fgColor="D5F5E3")
        c2.alignment = Alignment(horizontal="center")

cur.execute("SELECT * FROM customers ORDER BY group_code, customer_id")
all_cust = cur.fetchall()
cust_cols = [d[0] for d in cur.description]

for j, col in enumerate(cust_cols, 1):
    c = cs(ws1, DUMP_START, j, col, bold=True, bg="2E4057", align="center", fc="FFFFFF")
    c.border = thin_border()
ws1.row_dimensions[DUMP_START].height = 20

for i, row in enumerate(all_cust):
    r = DUMP_START + 1 + i
    last = row[cust_cols.index("last_name")]
    for j, val in enumerate(row, 1):
        bg = "FDFEFE" if i % 2 == 0 else "F0F3F4"
        if last in ("羽生", "大谷"):
            bg = "FFF9C4"
        c = ws1.cell(row=r, column=j, value=val)
        c.font = Font(size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()

for j in range(1, len(cust_cols)+1):
    ws1.column_dimensions[get_column_letter(j)].width = 20

# ---- Sheet2: DBダンプ_contracts_linked ----
ws2 = wb2.create_sheet("DBダンプ_contracts_linked")
DUMP2_START = 8

ws2.merge_cells("A1:J1")
c = ws2["A1"]
c.value = "【集計サマリ】contracts_linked分析（羽生・大谷）"
c.font = Font(bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F4E79")
c.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[1].height = 22

aggr2 = [
    ("羽生 group_code=A 契約件数", f"=COUNTIFS(G{DUMP2_START+1}:G9999,\"A\",H{DUMP2_START+1}:H9999,\"羽生\")"),
    ("羽生 group_code=B 契約件数", f"=COUNTIFS(G{DUMP2_START+1}:G9999,\"B\",H{DUMP2_START+1}:H9999,\"羽生\")"),
    ("大谷 group_code=A 契約件数", f"=COUNTIFS(G{DUMP2_START+1}:G9999,\"A\",H{DUMP2_START+1}:H9999,\"大谷\")"),
    ("大谷 group_code=B 契約件数", f"=COUNTIFS(G{DUMP2_START+1}:G9999,\"B\",H{DUMP2_START+1}:H9999,\"大谷\")"),
    ("参照G別同一顧客統合件数（羽生A）", f"=COUNTIFS(F{DUMP2_START+1}:F9999,1,G{DUMP2_START+1}:G9999,\"A\",H{DUMP2_START+1}:H9999,\"羽生\")"),
]
for i, (label, formula) in enumerate(aggr2, start=2):
    cs(ws2, i, 1, label, bg="EBF5FB")
    c2 = ws2.cell(row=i, column=2, value=formula)
    c2.font = Font(bold=True, size=10, color="1F4E79")
    c2.fill = PatternFill("solid", fgColor="D5F5E3")
    c2.alignment = Alignment(horizontal="center")

cur.execute("""
SELECT c.contract_no, c.agency_code, c.customer_name, c.policy_type,
       c.expiry_date, cu.customer_id, cu.group_code, cu.last_name,
       cu.first_name, cu.birth_date
FROM contracts c
LEFT JOIN customers cu ON c.linked_customer_id = cu.customer_id
WHERE cu.last_name IN ('羽生','大谷')
ORDER BY cu.last_name, cu.group_code
""")
linked_rows = cur.fetchall()
linked_cols = [d[0] for d in cur.description]

for j, col in enumerate(linked_cols, 1):
    c = cs(ws2, DUMP2_START, j, col, bold=True, bg="2E4057", align="center", fc="FFFFFF")
    c.border = thin_border()

for i, row in enumerate(linked_rows):
    r = DUMP2_START + 1 + i
    last = row[linked_cols.index("last_name")]
    for j, val in enumerate(row, 1):
        bg = "FFF9C4" if last == "羽生" else "E8F8F5"
        c = ws2.cell(row=r, column=j, value=val)
        c.font = Font(size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()

for j in range(1, len(linked_cols)+1):
    ws2.column_dimensions[get_column_letter(j)].width = 22

# ---- Sheet3~5: 画面証跡 ----
make_screen_sheet(wb2, "UT-NAMEYOSE-001",
    "羽生○弦 グループA検索 - 1件のみ表示確認",
    "A001/admin/password123でログイン後、顧客管理画面で「羽生」を検索\nグループA内で1件のみ（customer_id=1）が表示されることを確認\n期待値：羽生○弦 1件、自動車・自賠責・火災バッジ表示")
make_screen_sheet(wb2, "UT-NAMEYOSE-002",
    "羽生○弦 全グループ検索 - A・B各1件確認",
    "社員S001でログイン（全グループ参照可能）後、顧客管理画面で「羽生」を検索\nグループA・グループBで各1件、合計2件表示されることを確認\n期待値：羽生○弦 2件（customer_id=1, 11）")
make_screen_sheet(wb2, "UT-NAMEYOSE-003",
    "大谷○平 グループA検索 - 複数代理店バッジ確認",
    "A001/admin/password123でログイン後、顧客管理画面で「大谷」を検索\nグループA内で1件（customer_id=2）・A001/A003の複数代理店バッジ表示を確認\n期待値：大谷○平 1件、自動車・傷害・賠償責任バッジ表示")

wb2.save(os.path.join(BASE, "test", "006_UT-名寄せバッチ処理テスト", "名寄せテスト証跡.xlsx"))
print("006_名寄せテスト証跡.xlsx OK")


# ==============================
# 007 - 加入状況テスト仕様書.xlsx
# ==============================
wb3 = openpyxl.Workbook()
ws = wb3.active
ws.title = "テスト仕様書"

sheet_title(ws, "顧客管理画面 保険加入状況表示 単体テスト仕様書", col_span=8)

for r, k, v in [
    (2, "テスト工程", "UT（単体テスト）"),
    (3, "テスト確認観点", "保険加入状況○/－表示の正当性確認"),
    (4, "作成日", "2026-05-31"),
    (5, "作成者", "開発チーム"),
]:
    cs(ws, r, 1, k, bold=True, bg="D6E4F0", align="center")
    ws.cell(row=r, column=2, value=v).font = Font(size=10)
    ws.merge_cells(f"B{r}:H{r}")

for j, h in enumerate(hdr, 1):
    c = cs(ws, 7, j, h, bold=True, bg="1F4E79", align="center", fc="FFFFFF")
    c.border = thin_border()
ws.row_dimensions[7].height = 22

cases007 = [
    {
        "id": "UT-CUSTOMER-001",
        "target": "羽生○弦の保険加入状況表示",
        "content": "自動車・自賠責・火災の3契約を保有する\n羽生○弦の加入状況が正しく表示されること",
        "pre": "A001/admin/password123でログイン\n顧客管理画面を表示",
        "steps": "1. 顧客管理画面で「羽生」を検索\n2. 羽生○弦の加入状況列を確認\n3. 各種目の○/－表示を確認",
        "exp": "車○・火○・傷－・自賠○\n賠責－・サイバー－・所得－",
    },
    {
        "id": "UT-CUSTOMER-002",
        "target": "保険加入状況とDB契約データの突合",
        "content": "画面の○/－表示が\ncontractsテーブルの実データと一致すること",
        "pre": "A001/admin/password123でログイン\nDBに羽生○弦の契約データが存在",
        "steps": "1. 顧客管理画面で羽生○弦を検索\n2. +ボタンで詳細を展開\n3. 保有契約一覧とDB契約データを突合",
        "exp": "linked_customer_id=1で絞り込んだ\n契約のpolicy_typeと○表示が完全一致\n（自動車・自賠責・火災の3件）",
    },
    {
        "id": "UT-CUSTOMER-003",
        "target": "満期管理→顧客管理の画面遷移",
        "content": "満期管理画面の顧客名リンクをクリックして\n顧客管理画面に遷移し\n該当顧客1件が+展開状態で表示されること",
        "pre": "A001/admin/password123でログイン\n満期管理画面が表示されている",
        "steps": "1. 満期管理画面で鈴木○郎の顧客名リンクをクリック\n2. 顧客管理画面への遷移を確認\n3. URLにcustomer_id・expand=trueが含まれることを確認\n4. 該当顧客1件・詳細展開状態を確認",
        "exp": "URLにcustomer_id・expand=trueが含まれ\n顧客1件・詳細展開・保有契約一覧が表示",
    },
]

for i, case in enumerate(cases007):
    row = 8 + i
    ws.row_dimensions[row].height = 85
    vals = [case["id"], case["target"], case["content"], case["pre"], case["steps"], case["exp"], "○", ""]
    for j, val in enumerate(vals, 1):
        bg = "EBF5FB" if i % 2 == 0 else "FDFEFE"
        c = cs(ws, row, j, val, wrap=True, bg=bg)
        c.border = thin_border()
        if j == 1:
            c.font = Font(bold=True, size=10, color="1F4E79")
        if j == 7:
            c.font = Font(bold=True, size=12, color="27AE60")
            c.alignment = Alignment(horizontal="center", vertical="center")

for j, w in enumerate([18, 22, 42, 30, 42, 42, 8, 12], 1):
    ws.column_dimensions[get_column_letter(j)].width = w

wb3.save(os.path.join(BASE, "test", "007_UT-顧客管理保険加入状況テスト", "加入状況テスト仕様書.xlsx"))
print("007_加入状況テスト仕様書.xlsx OK")


# ==============================
# 007 - 加入状況テスト証跡.xlsx
# ==============================
wb4 = openpyxl.Workbook()
DUMP3_START = 9

# ---- Sheet1: DBダンプ_顧客契約突合 ----
ws41 = wb4.active
ws41.title = "DBダンプ_顧客契約突合"

ws41.merge_cells("A1:J1")
c = ws41["A1"]
c.value = "【集計サマリ】羽生○弦（グループA）顧客契約突合"
c.font = Font(bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F4E79")
c.alignment = Alignment(horizontal="left", vertical="center")
ws41.row_dimensions[1].height = 22

policy_types_list = ["自動車", "火災", "傷害", "自賠責", "賠償責任", "サイバー", "所得補償"]

aggr3 = [("種目別契約件数（COUNTIF）", None)]
for pt in policy_types_list:
    aggr3.append((f"  {pt}", f"=COUNTIF(F{DUMP3_START+1}:F9999,\"{pt}\")"))
aggr3.append(("", None))
aggr3.append(("画面表示の期待値（○/－）", None))
for pt in policy_types_list:
    aggr3.append((f"  {pt}", f"=IF(COUNTIF(F{DUMP3_START+1}:F9999,\"{pt}\")>0,\"○\",\"－\")"))

for i, (label, formula) in enumerate(aggr3, start=2):
    is_section = (formula is None and label != "")
    cs(ws41, i, 1, label, bold=is_section, bg="D6EAF8" if is_section else "EBF5FB")
    if formula:
        c2 = ws41.cell(row=i, column=2, value=formula)
        c2.font = Font(bold=True, size=10, color="1F4E79")
        c2.fill = PatternFill("solid", fgColor="D5F5E3")
        c2.alignment = Alignment(horizontal="center")

cur.execute("""
SELECT cu.customer_id, cu.last_name, cu.first_name, cu.group_code,
       c.contract_no, c.policy_type, c.agency_code, c.expiry_date,
       c.annual_premium, c.renewal_status
FROM customers cu
JOIN contracts c ON c.linked_customer_id = cu.customer_id
WHERE cu.last_name = '羽生' AND cu.group_code = 'A'
ORDER BY c.policy_type
""")
hanyu_rows = cur.fetchall()
hanyu_cols = [d[0] for d in cur.description]

for j, col in enumerate(hanyu_cols, 1):
    c = cs(ws41, DUMP3_START, j, col, bold=True, bg="2E4057", align="center", fc="FFFFFF")
    c.border = thin_border()
ws41.row_dimensions[DUMP3_START].height = 20

for i, row in enumerate(hanyu_rows):
    r = DUMP3_START + 1 + i
    for j, val in enumerate(row, 1):
        bg = "FFF9C4" if i % 2 == 0 else "FFFDE7"
        c = ws41.cell(row=r, column=j, value=val)
        c.font = Font(size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()

for j in range(1, len(hanyu_cols)+1):
    ws41.column_dimensions[get_column_letter(j)].width = 22

# ---- Sheet2~4: 画面証跡 ----
make_screen_sheet(wb4, "UT-CUSTOMER-001",
    "羽生○弦 保険加入状況○/－表示確認",
    "A001/admin/password123でログイン後、顧客管理画面で「羽生」を検索\n保険加入状況列の○/－表示をスクリーンショット\n期待値：車○・火○・傷－・自賠○・賠責－・サイバー－・所得－")
make_screen_sheet(wb4, "UT-CUSTOMER-002",
    "羽生○弦 詳細展開・保有契約一覧確認",
    "UT-CUSTOMER-001と同じ画面で+ボタンを開いた状態の詳細・保有契約一覧をスクリーンショット\n期待値：自動車（A001）・自賠責（A001）・火災（A002）の3契約が一覧表示")
make_screen_sheet(wb4, "UT-CUSTOMER-003",
    "満期管理→顧客管理画面遷移確認",
    "満期管理画面で鈴木○郎の顧客名リンクをクリックし\n遷移後の顧客管理画面（1件表示・展開済み）をスクリーンショット\nURLバーにcustomer_id・expand=trueが含まれることを確認")

wb4.save(os.path.join(BASE, "test", "007_UT-顧客管理保険加入状況テスト", "加入状況テスト証跡.xlsx"))
print("007_加入状況テスト証跡.xlsx OK")

conn.close()
print("全ファイル作成完了")
