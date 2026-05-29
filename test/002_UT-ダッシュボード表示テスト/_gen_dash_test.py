"""
ダッシュボードテスト仕様書・証跡 生成スクリプト
  - ダッシュボードテスト仕様書.xlsx（Sheet1：テスト仕様書 25件）
  - ダッシュボードテスト証跡.xlsx（Sheet1：DBダンプ_contracts / Sheet2：DBダンプ_users / Sheet3-5：画面キャプチャ）
"""
import os, sqlite3, time
from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from playwright.sync_api import sync_playwright

# ── パス定数 ──────────────────────────────────────────────────────
BASE_DIR  = Path(__file__).parent.parent.parent
DB_PATH   = str(BASE_DIR / "db" / "users.sqlite")
TEST_DIR  = Path(__file__).parent
SS_DIR    = TEST_DIR / "screenshots"
SPEC_PATH = TEST_DIR / "ダッシュボードテスト仕様書.xlsx"
EVID_PATH = TEST_DIR / "ダッシュボードテスト証跡.xlsx"
LOGIN_URL = (BASE_DIR / "frontend" / "login.html").as_uri()
TODAY     = date.today().strftime("%Y/%m/%d")

SS_DIR.mkdir(exist_ok=True)

# ── スタイル ──────────────────────────────────────────────────────
NAVY  = "0D2B5E"
GOLD  = "C9A84C"
WHITE = "FFFFFF"
LIGHT = "EEF2FA"
GREEN = "1E7E34"
AMBER = "FFF8E6"

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(c):   return PatternFill("solid", fgColor=c)
def center():  return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():    return Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── テストケース定義（25件）────────────────────────────────────────
# (明細番号, 確認対象画面, オペレーション内容, 期待結果)
SPEC_CASES = [
    # ── 固定表示確認 ──
    ("UT-DASH-001", "dashboard.html",
     "ページを開いてヘッダーロゴエリアを目視確認する",
     "会社名「AX損害保険株式会社」がヘッダー左に金色テキストで表示されること"),
    ("UT-DASH-002", "dashboard.html",
     "ページを開いてヘッダーロゴエリアを目視確認する",
     "サブタイトル「代理店Webシステム」がヘッダー左に白テキストで表示されること"),
    ("UT-DASH-003", "dashboard.html",
     "ページを開いてヘッダーロゴエリアを目視確認する",
     "盾マークのSVGロゴがヘッダー左端に金色で表示されること"),
    ("UT-DASH-004", "dashboard.html",
     "ページを開いてメニューエリアを目視確認する",
     "「顧客管理システム」ボタン（人物アイコン付き）が表示されること"),
    ("UT-DASH-005", "dashboard.html",
     "ページを開いてメニューエリアを目視確認する",
     "「満期管理システム」ボタン（時計アイコン付き）が表示されること"),
    ("UT-DASH-006", "dashboard.html",
     "ページを開いてメニューエリアを目視確認する",
     "「契約照会システム」ボタン（書類アイコン付き）が表示されること"),
    ("UT-DASH-007", "dashboard.html",
     "ページを開いてメニューエリアを目視確認する",
     "「保険金支払状況」ボタン（円マークアイコン付き）が表示されること"),
    ("UT-DASH-008", "dashboard.html",
     "ページを開いてインフォメーションエリアを目視確認する",
     "「保険会社からのお知らせ」欄にサンプルお知らせ3件が表示されること"),
    ("UT-DASH-009", "dashboard.html",
     "ページを開いてインフォメーションエリアを目視確認する",
     "「TODOタスク」小窓に進捗バーとタスクリストが表示されること"),
    ("UT-DASH-010", "dashboard.html",
     "ページを開いてインフォメーションエリアを目視確認する",
     "「お客様コンタクト履歴」欄にサンプル履歴3件が表示されること"),
    # ── ログインユーザー連動確認 ──
    ("UT-DASH-011", "dashboard.html",
     "A001/admin/password123でログインしてダッシュボードを表示する",
     "ヘッダー右上に代理店コード「A001」が表示されること"),
    ("UT-DASH-012", "dashboard.html",
     "A001/admin/password123でログインしてダッシュボードを表示する",
     "ヘッダー右上にログインID「admin」が表示されること"),
    ("UT-DASH-013", "dashboard.html",
     "A001/admin/password123でログインしてダッシュボードを表示する",
     "「ログアウト」ボタンがヘッダー右上に表示されること"),
    ("UT-DASH-014", "dashboard.html",
     "A001/admin/password123でログインしてダッシュボードを表示する",
     "今月更改グラフ中央に完了件数「28」が表示されること"),
    ("UT-DASH-015", "dashboard.html",
     "A001/admin/password123でログインしてダッシュボードを表示する",
     "翌月更改グラフ中央に完了件数「125」が表示されること"),
    ("UT-DASH-016", "dashboard.html",
     "A001/admin/password123でログインしてダッシュボードを表示する",
     "今月更改グラフ内バッジに完了率「11.4%」が表示されること"),
    ("UT-DASH-017", "dashboard.html",
     "A001/admin/password123でログインしてダッシュボードを表示する",
     "翌月更改グラフ内バッジに完了率「38.6%」が表示されること"),
    ("UT-DASH-018", "dashboard.html",
     "B002/agent1/pass456でログインしてダッシュボードを表示する",
     "B002の契約データ（今月15件完了/195件・翌月80件完了/290件）が表示されること"),
    ("UT-DASH-019", "dashboard.html",
     "C003/user1/pass789でログインしてダッシュボードを表示する",
     "C003の契約データ（今月10件完了/155件・翌月60件完了/240件）が表示されること"),
    # ── インフォグラフィックス集計正確性確認 ──
    ("UT-DASH-020", "dashboard.html / GET /api/dashboard",
     "A001でログインしてAPI レスポンス current_month.total を確認する",
     "A001今月総件数が DB の COUNTIFS 結果「245」と一致すること"),
    ("UT-DASH-021", "dashboard.html / GET /api/dashboard",
     "A001でログインしてAPI レスポンス current_month.completed を確認する",
     "A001今月更改済件数が DB の COUNTIFS 結果「28」と一致すること"),
    ("UT-DASH-022", "dashboard.html / GET /api/dashboard",
     "A001でログインしてAPI レスポンス current_month.pending を確認する",
     "A001今月未更改件数が DB の COUNTIFS 結果「217」と一致すること"),
    ("UT-DASH-023", "dashboard.html / GET /api/dashboard",
     "A001でログインしてAPI レスポンス next_month.total を確認する",
     "A001翌月総件数が DB の COUNTIFS 結果「324」と一致すること"),
    ("UT-DASH-024", "dashboard.html / GET /api/dashboard",
     "B002/agent1/pass456でログインしてAPI レスポンスを確認する",
     "B002の集計値（今月195件/完了15/未更改180・翌月290件/完了80/未更改210）が DB 値と一致すること"),
    ("UT-DASH-025", "dashboard.html / GET /api/dashboard",
     "C003/user1/pass789でログインしてAPI レスポンスを確認する",
     "C003の集計値（今月155件/完了10/未更改145・翌月240件/完了60/未更改180）が DB 値と一致すること"),
]
COLUMNS = ["明細番号", "確認対象画面", "オペレーション内容", "期待結果", "テスト結果", "テスト実施日", "テスト確認日"]


# ════════════════════════════════════════════════════════
# 1. テスト仕様書生成
# ════════════════════════════════════════════════════════
def gen_spec():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "テスト仕様書"
    LAST = get_column_letter(len(COLUMNS))

    # 行1: タイトル
    ws.merge_cells(f"A1:{LAST}1")
    c = ws["A1"]
    c.value = "dashboard.html　ダッシュボード表示機能　単体テスト仕様書"
    c.font      = Font(name="游ゴシック", bold=True, size=14, color=WHITE)
    c.fill      = fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 30

    # 行2: メタ情報
    ws.merge_cells("A2:C2")
    m1 = ws["A2"]
    m1.value     = f"テストID：UT-DASH-001 ～ UT-DASH-{len(SPEC_CASES):03d}　　テスト確認観点：全条件網羅テスト"
    m1.font      = Font(name="游ゴシック", bold=True, size=10, color=NAVY)
    m1.fill      = fill(AMBER)
    m1.alignment = left()
    ws.merge_cells(f"D2:{LAST}2")
    m2 = ws["D2"]
    m2.value     = f"テスト工程：UT（単体テスト）　　作成日：{TODAY}"
    m2.font      = Font(name="游ゴシック", bold=True, size=10, color=NAVY)
    m2.fill      = fill(AMBER)
    m2.alignment = left()
    ws.row_dimensions[2].height = 22

    # 行3: ゴールドライン
    for col in range(1, len(COLUMNS) + 1):
        ws.cell(row=3, column=col).fill = fill(GOLD)
    ws.row_dimensions[3].height = 4

    # 行4: ヘッダー
    for ci, name in enumerate(COLUMNS, 1):
        c = ws.cell(row=4, column=ci, value=name)
        c.font      = Font(name="游ゴシック", bold=True, size=10, color=WHITE)
        c.fill      = fill(NAVY)
        c.alignment = center()
        c.border    = thin_border()
    ws.row_dimensions[4].height = 24

    # 行5〜: データ
    groups = [
        ("固定表示確認",               10),
        ("ログインユーザー連動確認",     9),
        ("インフォグラフィックス集計正確性確認", 6),
    ]
    group_idx = 0; group_count = 0; group_start = 5
    for ri, case in enumerate(SPEC_CASES, 5):
        bg = WHITE if ri % 2 == 1 else LIGHT
        row_data = list(case) + ["", "", ""]  # テスト結果・実施日・確認日は空欄
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border    = thin_border()
            c.fill      = fill(bg)
            c.font      = Font(name="游ゴシック", size=10, color="1A1A2E")
            c.alignment = center() if ci in (1, 2, 5, 6, 7) else left()
        ws.row_dimensions[ri].height = 38

    # グループ区切り行（背景色で区別）
    group_rows = {5: "固定表示確認（001〜010）", 15: "ログインユーザー連動確認（011〜019）", 24: "インフォグラフィックス集計正確性確認（020〜025）"}
    for gr, label in group_rows.items():
        ws.insert_rows(gr)
        ws.merge_cells(f"A{gr}:{LAST}{gr}")
        gc = ws[f"A{gr}"]
        gc.value     = f"■ {label}"
        gc.font      = Font(name="游ゴシック", bold=True, size=9, color=WHITE)
        gc.fill      = fill("2D4F8E")
        gc.alignment = left()
        ws.row_dimensions[gr].height = 18

    # 列幅
    for i, w in enumerate([18, 28, 52, 50, 12, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows = "1:4"

    wb.save(str(SPEC_PATH))
    print(f"仕様書生成完了: {SPEC_PATH}")


# ════════════════════════════════════════════════════════
# 2. Playwright でスクリーンショット取得
# ════════════════════════════════════════════════════════
ACCOUNTS = [
    ("UT-DASH-011", "A001", "admin",  "password123"),
    ("UT-DASH-018", "B002", "agent1", "pass456"),
    ("UT-DASH-019", "C003", "user1",  "pass789"),
]

def take_screenshots() -> dict:
    """各アカウントでログインしてダッシュボードのスクリーンショットを保存する"""
    paths = {}
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx     = browser.new_context(viewport={"width": 1400, "height": 860})
        page    = ctx.new_page()

        for test_id, agency, login_id, password in ACCOUNTS:
            print(f"  スクリーンショット取得: {test_id} ({agency}/{login_id})")
            page.goto(LOGIN_URL)
            page.wait_for_load_state("domcontentloaded")
            # localStorageをクリアして前回ログイン情報をリセットする
            page.evaluate("localStorage.clear()")
            page.goto(LOGIN_URL)
            page.wait_for_load_state("domcontentloaded")
            page.fill("#agency-code", agency)
            page.fill("#login-id",    login_id)
            page.fill("#password",    password)
            page.click("button[type='submit']")
            page.wait_for_url("**/dashboard.html", timeout=8000)
            time.sleep(2.5)
            ss = str(SS_DIR / f"{test_id}.png")
            page.screenshot(path=ss)
            paths[test_id] = ss

        browser.close()
    return paths


# ════════════════════════════════════════════════════════
# 3. テスト証跡生成
# ════════════════════════════════════════════════════════
def gen_evidence(ss_paths: dict):
    wb  = openpyxl.Workbook()

    # ── Sheet1: DBダンプ_contracts ──────────────────────
    ws1 = wb.active
    ws1.title = "DBダンプ_contracts"

    # DB から contracts 全件取得
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()
    cur.execute("SELECT id, agency_code, contract_no, customer_name, renewal_month, status, created_at FROM contracts ORDER BY agency_code, renewal_month, status")
    contracts = cur.fetchall()
    total_rows = len(contracts)
    # データ開始行は行12（行11がヘッダー）
    DATA_START = 12
    DATA_END   = DATA_START + total_rows - 1
    # 数式用の参照範囲
    B = f"$B${DATA_START}:$B${DATA_END}"
    E = f"$E${DATA_START}:$E${DATA_END}"
    F = f"$F${DATA_START}:$F${DATA_END}"

    # ── タイトル行 ──
    ws1.merge_cells("A1:G1")
    t = ws1["A1"]
    t.value     = "DBダンプ：contracts テーブル 集計マトリクス（COUNTIFS）"
    t.font      = Font(name="游ゴシック", bold=True, size=13, color=WHITE)
    t.fill      = fill(NAVY)
    t.alignment = center()
    ws1.row_dimensions[1].height = 28

    # ── 行2〜3: 集計マトリクスヘッダー ──
    ws1.merge_cells("A2:G2")
    ws1["A2"].value     = "集計マトリクス（代理店コード × 月別 × ステータス）"
    ws1["A2"].font      = Font(name="游ゴシック", bold=True, size=10, color=NAVY)
    ws1["A2"].fill      = fill(AMBER)
    ws1["A2"].alignment = left()
    ws1.row_dimensions[2].height = 20

    matrix_headers = ["代理店コード",
                      "2026-05 総件数", "2026-05 更改済", "2026-05 未更改",
                      "2026-06 総件数", "2026-06 更改済", "2026-06 未更改"]
    for ci, h in enumerate(matrix_headers, 1):
        c = ws1.cell(row=3, column=ci, value=h)
        c.font      = Font(name="游ゴシック", bold=True, size=9, color=WHITE)
        c.fill      = fill(NAVY)
        c.alignment = center()
        c.border    = thin_border()
    ws1.row_dimensions[3].height = 20

    # ── 行4〜6: 各代理店の COUNTIFS 数式 ──
    agencies = ["A001", "B002", "C003"]
    for ri, ag in enumerate(agencies, 4):
        bg = WHITE if ri % 2 == 1 else LIGHT
        ws1.cell(row=ri, column=1, value=ag).font = Font(name="游ゴシック", bold=True, size=10, color=NAVY)
        ws1.cell(row=ri, column=1).fill      = fill(bg)
        ws1.cell(row=ri, column=1).alignment = center()
        ws1.cell(row=ri, column=1).border    = thin_border()
        # 今月
        ws1.cell(row=ri, column=2, value=f'=COUNTIFS({B},"{ag}",{E},"2026-05")')
        ws1.cell(row=ri, column=3, value=f'=COUNTIFS({B},"{ag}",{E},"2026-05",{F},"completed")')
        ws1.cell(row=ri, column=4, value=f'=COUNTIFS({B},"{ag}",{E},"2026-05",{F},"pending")')
        # 翌月
        ws1.cell(row=ri, column=5, value=f'=COUNTIFS({B},"{ag}",{E},"2026-06")')
        ws1.cell(row=ri, column=6, value=f'=COUNTIFS({B},"{ag}",{E},"2026-06",{F},"completed")')
        ws1.cell(row=ri, column=7, value=f'=COUNTIFS({B},"{ag}",{E},"2026-06",{F},"pending")')
        for ci in range(2, 8):
            ws1.cell(row=ri, column=ci).font      = Font(name="游ゴシック", size=10)
            ws1.cell(row=ri, column=ci).fill      = fill(bg)
            ws1.cell(row=ri, column=ci).alignment = center()
            ws1.cell(row=ri, column=ci).border    = thin_border()
        ws1.row_dimensions[ri].height = 18

    # ── 行7: 合計 ──
    ws1.cell(row=7, column=1, value="合計").font = Font(name="游ゴシック", bold=True, size=10, color=WHITE)
    ws1.cell(row=7, column=1).fill      = fill(NAVY)
    ws1.cell(row=7, column=1).alignment = center()
    ws1.cell(row=7, column=1).border    = thin_border()
    for ci in range(2, 8):
        ws1.cell(row=7, column=ci, value=f"=SUM({get_column_letter(ci)}4:{get_column_letter(ci)}6)")
        ws1.cell(row=7, column=ci).font      = Font(name="游ゴシック", bold=True, size=10, color=WHITE)
        ws1.cell(row=7, column=ci).fill      = fill(NAVY)
        ws1.cell(row=7, column=ci).alignment = center()
        ws1.cell(row=7, column=ci).border    = thin_border()
    ws1.row_dimensions[7].height = 18

    # ── 行8: 空白・区切り ──
    ws1.row_dimensions[8].height = 8
    ws1.merge_cells("A9:G9")
    ws1["A9"].value     = f"全レコード一覧（{total_rows}件）"
    ws1["A9"].font      = Font(name="游ゴシック", bold=True, size=10, color=WHITE)
    ws1["A9"].fill      = fill("2D4F8E")
    ws1["A9"].alignment = left()
    ws1.row_dimensions[9].height = 18

    # ── 行10: 空白 ──
    ws1.row_dimensions[10].height = 4

    # ── 行11: データヘッダー ──
    data_cols = ["id", "agency_code", "contract_no", "customer_name", "renewal_month", "status", "created_at"]
    for ci, h in enumerate(data_cols, 1):
        c = ws1.cell(row=11, column=ci, value=h)
        c.font      = Font(name="游ゴシック", bold=True, size=9, color=WHITE)
        c.fill      = fill(NAVY)
        c.alignment = center()
        c.border    = thin_border()
    ws1.row_dimensions[11].height = 18

    # ── 行12〜: 全レコード ──
    for ri, row in enumerate(contracts, DATA_START):
        bg = WHITE if ri % 2 == 0 else LIGHT
        for ci, val in enumerate(row, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="游ゴシック", size=8)
            c.fill      = fill(bg)
            c.alignment = center() if ci in (1, 2, 5, 6) else left()
            c.border    = thin_border()
        ws1.row_dimensions[ri].height = 14

    # 列幅
    for i, w in enumerate([6, 12, 22, 18, 12, 12, 18], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet2: DBダンプ_users ──────────────────────────
    ws2 = wb.create_sheet("DBダンプ_users")

    cur.execute("SELECT id, agency_code, login_id, name, created_at FROM users ORDER BY id")
    users = cur.fetchall()
    conn.close()

    ws2.merge_cells("A1:E1")
    t2 = ws2["A1"]
    t2.value     = "DBダンプ：users テーブル（代理店コード・ログインID 証跡）"
    t2.font      = Font(name="游ゴシック", bold=True, size=13, color=WHITE)
    t2.fill      = fill(NAVY)
    t2.alignment = center()
    ws2.row_dimensions[1].height = 28

    user_cols = ["id", "agency_code", "login_id", "name", "created_at"]
    for ci, h in enumerate(user_cols, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font      = Font(name="游ゴシック", bold=True, size=10, color=WHITE)
        c.fill      = fill(NAVY)
        c.alignment = center()
        c.border    = thin_border()
    ws2.row_dimensions[2].height = 22

    for ri, row in enumerate(users, 3):
        for ci, val in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="游ゴシック", size=10)
            c.fill      = fill(WHITE if ri % 2 == 1 else LIGHT)
            c.alignment = center() if ci in (1, 2, 3) else left()
            c.border    = thin_border()
        ws2.row_dimensions[ri].height = 20

    for i, w in enumerate([6, 14, 14, 18, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet3〜5: 画面キャプチャ証跡 ───────────────────
    for test_id, img_path in ss_paths.items():
        ws = wb.create_sheet(test_id)

        ws.merge_cells("A1:H1")
        h = ws["A1"]
        h.value     = f"{test_id}　画面キャプチャ証跡"
        h.font      = Font(name="游ゴシック", bold=True, size=12, color=WHITE)
        h.fill      = fill(NAVY)
        h.alignment = center()
        ws.row_dimensions[1].height = 28

        if os.path.exists(img_path):
            img        = XlImage(img_path)
            img.width  = 900
            img.height = 563
            ws.add_image(img, "A3")
            ws.row_dimensions[3].height = 430
        ws.column_dimensions["A"].width = 18

    wb.save(str(EVID_PATH))
    print(f"証跡生成完了: {EVID_PATH}")


# ════════════════════════════════════════════════════════
# メイン
# ════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 55)
    print("  ダッシュボードテスト 仕様書・証跡 生成")
    print("=" * 55)

    print("\n【Step 1】テスト仕様書 生成")
    gen_spec()

    print("\n【Step 2】Playwright スクリーンショット取得")
    ss_paths = take_screenshots()
    print(f"  取得完了: {len(ss_paths)}件")

    print("\n【Step 3】テスト証跡 生成")
    gen_evidence(ss_paths)

    print("\n" + "=" * 55)
    print("  完了")
    print("=" * 55)
