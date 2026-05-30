"""
権限管理テスト証跡.xlsx に UT-AUTH-001〜017 の個別シートを追加し、
UT-003 と同一フォーマットでスクリーンショットを貼り付けるスクリプト
"""
import io, pathlib
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL_PATH = pathlib.Path(__file__).parent / "権限管理テスト証跡.xlsx"
SS_DIR     = pathlib.Path(__file__).parent / "screenshots"

# ── UT-003 と同一の画像サイズ ─────────────────────────────────
IMG_W = 1400
IMG_H = 860

# ── テストケース定義 ──────────────────────────────────────────
# (シート名, 行1タイトル, 行2確認内容, 使用スクリーンショット)
CASES = [
    (
        "UT-AUTH-001",
        "UT-AUTH-001　ログイン画面表示確認",
        "確認内容：ブラウザで login.html を開き、ログイン画面（代理店コード・ログインID・パスワード入力欄・ログインボタン）が表示されることを確認",
        "01_login.png",
    ),
    (
        "UT-AUTH-002",
        "UT-AUTH-002　staff1 ログイン → ロール表示確認",
        "確認内容：A001/staff1 でログインし dashboard.html へ遷移。ヘッダーバッジに「A001 staff1｜一般担当」と表示されることを確認",
        "02_dashboard_staff1.png",
    ),
    (
        "UT-AUTH-003",
        "UT-AUTH-003　保険金支払状況ボタン グレーアウト確認",
        "確認内容：staff1（PAYMENT_VIEW権限なし）ログイン後、「保険金支払状況」ボタンがグレーアウト・クリック不可で、他3ボタンが活性状態であることを確認",
        "03_dashboard_staff1_permission.png",
    ),
    (
        "UT-AUTH-004",
        "UT-AUTH-004　権限なしボタン ツールチップ確認",
        "確認内容：グレーアウトされた「保険金支払状況」ボタンにホバーし「権限がありません」ツールチップが表示されることを確認",
        "04_tooltip_no_permission.png",
    ),
    (
        "UT-AUTH-005",
        "UT-AUTH-005　ユーザー管理リンク 非表示確認（staff1）",
        "確認内容：staff1 ログイン時、ヘッダー右上に「ユーザー管理」リンクが表示されない（display:none）ことを確認",
        "03_dashboard_staff1_permission.png",
    ),
    (
        "UT-AUTH-006",
        "UT-AUTH-006　ログアウト → ログイン画面復帰確認",
        "確認内容：staff1 ログイン後、「ログアウト」ボタンをクリックし localStorage がクリアされて login.html に戻ることを確認",
        "05_logout_back_to_login.png",
    ),
    (
        "UT-AUTH-007",
        "UT-AUTH-007　admin ログイン → 管理者ロール表示確認",
        "確認内容：A001/admin でログインし dashboard.html へ遷移。ヘッダーバッジに「A001 admin｜管理者」と表示されることを確認",
        "06_dashboard_admin.png",
    ),
    (
        "UT-AUTH-008",
        "UT-AUTH-008　全メニューボタン 活性確認（admin）",
        "確認内容：admin ログイン後、顧客管理・満期管理・契約照会・保険金支払状況 の全4ボタンがグレーアウトなし・活性状態であることを確認",
        "07_dashboard_admin_permission.png",
    ),
    (
        "UT-AUTH-009",
        "UT-AUTH-009　ユーザー管理リンク 表示確認（admin）",
        "確認内容：admin ログイン後、ヘッダー右上に「ユーザー管理」リンクが表示されることを確認",
        "07_dashboard_admin_permission.png",
    ),
    (
        "UT-AUTH-010",
        "UT-AUTH-010　ユーザー管理リンク → admin.html 遷移確認",
        "確認内容：admin ログイン後、ヘッダーの「ユーザー管理」リンクをクリックし admin.html へ遷移・ユーザー管理画面が表示されることを確認",
        "08_admin_page.png",
    ),
    (
        "UT-AUTH-011",
        "UT-AUTH-011　ユーザー一覧テーブル 表示確認",
        "確認内容：ユーザー管理画面にて admin（管理者）・staff1（一般担当）の2件がロールチップ・有効バッジ・編集/削除ボタンとともに表示されることを確認",
        "09_admin_user_list.png",
    ),
    (
        "UT-AUTH-012",
        "UT-AUTH-012　ユーザー追加モーダル 表示確認",
        "確認内容：「＋ユーザー追加」ボタンをクリックし、ログインID・パスワード・ロール・有効フラグの入力欄を持つ追加モーダルが表示されることを確認",
        "10_admin_add_modal.png",
    ),
    (
        "UT-AUTH-013",
        "UT-AUTH-013　ロール別権限マトリクス 表示確認",
        "確認内容：画面下部のロール別権限一覧（管理者=全権限✓、一般担当=保険金支払状況参照のみ—、閲覧専用=制限あり）が読み取り専用で表示されることを確認",
        "11_admin_permission_matrix.png",
    ),
    (
        "UT-AUTH-014",
        "UT-AUTH-014　GET /api/permissions (staff1) 権限リスト返却確認",
        "確認内容：staff1 の JWT で /api/permissions を呼び出し、role_id:2・role_name:一般担当・permissions:[CUSTOMER_EDIT, MATURITY_VIEW, REPORT_VIEW, USER_ADMIN] が返ることをダッシュボードの権限制御表示で確認",
        "03_dashboard_staff1_permission.png",
    ),
    (
        "UT-AUTH-015",
        "UT-AUTH-015　GET /api/users (staff1) → 403応答確認",
        "確認内容：staff1（role_id:2）で /api/users を呼び出すと HTTP 403 が返り、ユーザー管理リンク非表示・管理画面アクセス不可により確認",
        "03_dashboard_staff1_permission.png",
    ),
    (
        "UT-AUTH-016",
        "UT-AUTH-016　GET /api/users (admin) → ユーザー一覧返却確認",
        "確認内容：admin（role_id:1）で /api/users を呼び出し、A001 代理店のユーザー2件（admin・staff1）が返ることをユーザー一覧画面表示で確認",
        "09_admin_user_list.png",
    ),
    (
        "UT-AUTH-017",
        "UT-AUTH-017　非管理者の admin.html 直接アクセス → リダイレクト確認",
        "確認内容：staff1 ログイン状態で admin.html に直接 URL アクセスした際、role_id:2 と判定されて dashboard.html にリダイレクトされることを確認",
        "03_dashboard_staff1_permission.png",
    ),
]

# ── スタイル定数（UT-003 と同一）──────────────────────────────
NAVY     = "000D2B5E"
GOLD_BG  = "00FFF8E6"
WHITE    = "00FFFFFF"
GRAY_TXT = "005A6480"
NAVY_TXT = "000D2B5E"
EXEC_INFO = "テスト実施日：2026/05/30　　環境：ローカルファイル（file://）　　ブラウザ：Chromium (Playwright headless)"


def resize_to_buf(img_path: pathlib.Path, w: int, h: int) -> io.BytesIO:
    with PILImage.open(img_path) as im:
        im = im.convert("RGB")
        im = im.resize((w, h), PILImage.LANCZOS)
        buf = io.BytesIO()
        im.save(buf, format="PNG", optimize=True)
        buf.seek(0)
    return buf


def build_case_sheet(wb, sheet_name, row1_title, row2_content, img_path):
    ws = wb.create_sheet(sheet_name)

    # カラム幅（UT-003 準拠）
    ws.column_dimensions["A"].width = 18.0

    # ── 行1：タイトル（navy背景・白太字）────────────────────
    ws.row_dimensions[1].height = 28.0
    ws.merge_cells("A1:H1")
    c1 = ws["A1"]
    c1.value = row1_title
    c1.font  = Font(bold=True, size=12, color=WHITE, name="Meiryo")
    c1.fill  = PatternFill(fill_type="solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center")

    # ── 行2：確認内容（薄黄背景・紺文字）────────────────────
    ws.row_dimensions[2].height = 20.0
    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = row2_content
    c2.font  = Font(bold=False, size=10, color=NAVY_TXT, name="Meiryo")
    c2.fill  = PatternFill(fill_type="solid", fgColor=GOLD_BG)
    c2.alignment = Alignment(horizontal="left", vertical="center")

    # ── 行3：実施情報（グレー文字・背景なし）────────────────
    ws.row_dimensions[3].height = 16.0
    ws.merge_cells("A3:H3")
    c3 = ws["A3"]
    c3.value = EXEC_INFO
    c3.font  = Font(bold=False, size=9, color=GRAY_TXT, name="Meiryo")
    c3.alignment = Alignment(horizontal="left", vertical="center")

    # ── 行4〜：スクリーンショット（col=0, row=4 = A5セル相当）
    buf    = resize_to_buf(img_path, IMG_W, IMG_H)
    xl_img = XLImage(buf)
    xl_img.width  = IMG_W
    xl_img.height = IMG_H
    # UT-003 と同じアンカー：col=0, row=4 (0-indexed)
    ws.add_image(xl_img, "A5")

    return ws


def main():
    wb = load_workbook(EXCEL_PATH)

    # 前回追加した「スクリーンショット証跡」シートを削除
    if "スクリーンショット証跡" in wb.sheetnames:
        del wb["スクリーンショット証跡"]
        print("  既存の「スクリーンショット証跡」シートを削除")

    # 既存の個別ケースシートも削除（再実行対応）
    for case_id, *_ in CASES:
        if case_id in wb.sheetnames:
            del wb[case_id]

    # 各テストケースシートを追加
    for sheet_name, row1, row2, img_name in CASES:
        img_path = SS_DIR / img_name
        build_case_sheet(wb, sheet_name, row1, row2, img_path)
        print(f"  [{sheet_name}] ← {img_name}")

    wb.save(EXCEL_PATH)
    print(f"\n保存完了: {EXCEL_PATH}")
    print(f"シート構成: {wb.sheetnames}")


if __name__ == "__main__":
    main()
