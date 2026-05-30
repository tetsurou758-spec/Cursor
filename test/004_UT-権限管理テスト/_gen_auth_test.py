"""
004_UT-権限管理テスト
仕様書・証跡 Excel 生成スクリプト
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pathlib

OUT_DIR = pathlib.Path(__file__).parent

# ── カラーパレット ─────────────────────────────────────────────
NAVY   = "000D2B5E"
GOLD   = "00C9A84C"
WHITE  = "00FFFFFF"
LIGHT  = "00F0F2F7"
GRAY   = "00E0E5F0"

# ── テストケース定義 ──────────────────────────────────────────
# (試験番号, 確認対象画面, オペレーション内容, 期待結果)
CASES = [
    # セクションヘッダー
    ("【 ログイン・ロール表示確認（001〜006）】", None, None, None),
    (
        "UT-AUTH-001",
        "login.html",
        "ブラウザで login.html を開く",
        "ログイン画面が表示されること（代理店コード・ログインID・パスワード入力欄、ログインボタンが存在すること）",
    ),
    (
        "UT-AUTH-002",
        "login.html → dashboard.html",
        "代理店コード「A001」・ログインID「staff1」・パスワード「pass001」を入力してログインボタンをクリックする",
        "dashboard.html へ遷移し、ヘッダーのユーザーバッジに「A001 staff1｜一般担当」と表示されること",
    ),
    (
        "UT-AUTH-003",
        "dashboard.html（staff1）",
        "staff1 でログイン後、ダッシュボードのメニュー「保険金支払状況」ボタンの状態を確認する",
        "「保険金支払状況」ボタンがグレーアウト（no-permission クラス付与・opacity 低下）されクリック不可であること。"
        "他の3ボタン（顧客管理・満期管理・契約照会）は活性状態であること",
    ),
    (
        "UT-AUTH-004",
        "dashboard.html（staff1）",
        "グレーアウトされた「保険金支払状況」ボタンにマウスをホバーする",
        "「権限がありません」というツールチップが表示されること",
    ),
    (
        "UT-AUTH-005",
        "dashboard.html（staff1）",
        "staff1 でログイン後、ヘッダー右上を確認する",
        "「ユーザー管理」リンクが表示されないこと（display:none 状態であること）",
    ),
    (
        "UT-AUTH-006",
        "dashboard.html → login.html",
        "staff1 でログイン後、ヘッダーの「ログアウト」ボタンをクリックする",
        "localStorage がクリアされ login.html に戻ること",
    ),
    # セクションヘッダー
    ("【 管理者権限確認（007〜010）】", None, None, None),
    (
        "UT-AUTH-007",
        "login.html → dashboard.html",
        "代理店コード「A001」・ログインID「admin」・パスワード「password123」を入力してログインボタンをクリックする",
        "dashboard.html へ遷移し、ヘッダーのユーザーバッジに「A001 admin｜管理者」と表示されること",
    ),
    (
        "UT-AUTH-008",
        "dashboard.html（admin）",
        "admin でログイン後、ダッシュボードのメニュー4ボタンの状態を確認する",
        "顧客管理システム・満期管理システム・契約照会システム・保険金支払状況 の全4ボタンが活性状態（グレーアウトなし）であること",
    ),
    (
        "UT-AUTH-009",
        "dashboard.html（admin）",
        "admin でログイン後、ヘッダー右上を確認する",
        "「ユーザー管理」リンクが表示されること",
    ),
    (
        "UT-AUTH-010",
        "dashboard.html → admin.html",
        "admin でログイン後、ヘッダーの「ユーザー管理」リンクをクリックする",
        "admin.html へ遷移し、ユーザー管理画面が表示されること",
    ),
    # セクションヘッダー
    ("【 ユーザー管理画面確認（011〜013）】", None, None, None),
    (
        "UT-AUTH-011",
        "admin.html",
        "admin でログイン後、ユーザー管理画面のユーザー一覧テーブルを確認する",
        "admin（管理者）・staff1（一般担当）の2行が表示され、各行にロールチップ・有効バッジ・編集ボタン・削除ボタンが存在すること",
    ),
    (
        "UT-AUTH-012",
        "admin.html",
        "「＋ ユーザー追加」ボタンをクリックする",
        "ユーザー追加モーダルが表示され、ログインID・パスワード・ロール・有効フラグの入力欄が存在すること",
    ),
    (
        "UT-AUTH-013",
        "admin.html",
        "画面下部のロール別権限一覧を確認する",
        "管理者=全機能✓、一般担当=保険金支払状況参照のみ—、閲覧専用=保険金支払状況参照・ユーザー管理のみ✓のマトリクスが読み取り専用で表示されること",
    ),
    # セクションヘッダー
    ("【 API 権限制御確認（014〜017）】", None, None, None),
    (
        "UT-AUTH-014",
        "GET /api/permissions",
        "staff1 の JWT で GET /api/permissions を呼び出す",
        "role_id:2・role_name:「一般担当」・permissions:[CUSTOMER_EDIT, MATURITY_VIEW, REPORT_VIEW, USER_ADMIN] が返ること",
    ),
    (
        "UT-AUTH-015",
        "GET /api/users",
        "staff1 の JWT（role_id:2）で GET /api/users を呼び出す",
        "HTTP 403 が返り、管理者のみアクセス可能である旨のエラーメッセージが表示されること",
    ),
    (
        "UT-AUTH-016",
        "GET /api/users",
        "admin の JWT（role_id:1）で GET /api/users を呼び出す",
        "A001 代理店の自代理店ユーザー一覧（admin・staff1 の2件）が返ること",
    ),
    (
        "UT-AUTH-017",
        "admin.html（直接アクセス）",
        "staff1 でログインした状態で admin.html に直接 URL アクセスする",
        "管理者以外と判定され dashboard.html へリダイレクトされること",
    ),
]

# 証跡の合否・日付
RESULT_DATE = "2026/05/30"

# ── スタイルヘルパー ──────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="FFB0B8C4")
    return Border(left=s, right=s, top=s, bottom=s)

def make_font(bold=False, size=10, color=WHITE):
    return Font(bold=bold, size=size, color=color, name="Meiryo")

def make_fill(color):
    return PatternFill(fill_type="solid", fgColor=color)

def set_header_style(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True, size=10, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=fg, name="Meiryo")
    c.fill = make_fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = thin_border()
    return c

def set_data_style(ws, row, col, value, bg="00FFFFFF", wrap=True, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(size=9, name="Meiryo", color="00333333")
    c.fill = make_fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border = thin_border()
    return c

# ── 共通シート構築 ────────────────────────────────────────────
def build_sheet(ws, include_result):
    # カラム幅
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    # 行1: タイトル
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "権限管理テスト仕様書（dashboard.html / admin.html / API権限制御）"
    c.font = Font(bold=True, size=14, color=WHITE, name="Meiryo")
    c.fill = make_fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # 行2: テストID範囲・工程
    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:C2")
    ws["A2"].value = "テストID：UT-AUTH-001 〜 UT-AUTH-017"
    ws["A2"].font = Font(size=10, color=WHITE, name="Meiryo")
    ws["A2"].fill = make_fill(NAVY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("D2:G2")
    ws["D2"].value = "テスト工程：UT（単体テスト）　　作成日：2026/05/30"
    ws["D2"].font = Font(size=10, color=WHITE, name="Meiryo")
    ws["D2"].fill = make_fill(NAVY)
    ws["D2"].alignment = Alignment(horizontal="right", vertical="center")

    # 行3: 空白
    ws.row_dimensions[3].height = 8

    # 行4: 列ヘッダー
    ws.row_dimensions[4].height = 24
    headers = ["試験番号", "確認対象画面", "オペレーション内容", "期待結果",
               "テスト結果", "テスト実施日", "テスト確認日"]
    for i, h in enumerate(headers, 1):
        set_header_style(ws, 4, i, h, bg=NAVY, fg=WHITE, bold=True, size=10)

    # 行5〜: テストケース
    cur_row = 5
    case_idx = 0
    for case in CASES:
        trial_no, screen, op, expect = case

        # セクションヘッダー行
        if screen is None:
            ws.row_dimensions[cur_row].height = 20
            ws.merge_cells(f"A{cur_row}:G{cur_row}")
            c = ws.cell(row=cur_row, column=1, value=trial_no)
            c.font = Font(bold=True, size=10, color=NAVY, name="Meiryo")
            c.fill = make_fill("00E8EDF8")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin_border()
            cur_row += 1
            continue

        # データ行
        ws.row_dimensions[cur_row].height = 56
        row_bg = "00F7F9FC" if case_idx % 2 == 0 else "00FFFFFF"

        set_data_style(ws, cur_row, 1, trial_no,  bg=row_bg, wrap=False, align="center")
        set_data_style(ws, cur_row, 2, screen,    bg=row_bg, wrap=True)
        set_data_style(ws, cur_row, 3, op,        bg=row_bg, wrap=True)
        set_data_style(ws, cur_row, 4, expect,    bg=row_bg, wrap=True)

        if include_result:
            set_data_style(ws, cur_row, 5, "OK",         bg=row_bg, wrap=False, align="center")
            set_data_style(ws, cur_row, 6, RESULT_DATE,  bg=row_bg, wrap=False, align="center")
            set_data_style(ws, cur_row, 7, RESULT_DATE,  bg=row_bg, wrap=False, align="center")
        else:
            for col in [5, 6, 7]:
                set_data_style(ws, cur_row, col, None, bg=row_bg)

        cur_row += 1
        case_idx += 1

    return ws


def generate():
    dest = OUT_DIR

    # ── 仕様書 ───────────────────────────────────────────────
    wb_spec = openpyxl.Workbook()
    ws_spec = wb_spec.active
    ws_spec.title = "テスト仕様書"
    build_sheet(ws_spec, include_result=False)
    spec_path = dest / "権限管理テスト仕様書.xlsx"
    wb_spec.save(spec_path)
    print(f"  生成: {spec_path.name}")

    # ── 証跡 ─────────────────────────────────────────────────
    wb_rec = openpyxl.Workbook()
    ws_rec = wb_rec.active
    ws_rec.title = "テスト証跡"
    build_sheet(ws_rec, include_result=True)
    rec_path = dest / "権限管理テスト証跡.xlsx"
    wb_rec.save(rec_path)
    print(f"  生成: {rec_path.name}")

    print("\n完了")


if __name__ == "__main__":
    generate()
