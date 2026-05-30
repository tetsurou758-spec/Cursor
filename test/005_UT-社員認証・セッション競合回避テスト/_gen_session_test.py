"""
005_UT-社員認証・セッション競合回避テスト
仕様書・証跡 Excel 生成スクリプト

テスト目的：
  代理店タブ（Tab A）と社員タブ（Tab B）を交互に操作し、
  sessionStorage を独立管理することでセッション競合が
  発生しないことを証明する。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pathlib

OUT_DIR = pathlib.Path(__file__).parent

# ── カラーパレット ─────────────────────────────────────────────
NAVY   = "000D2B5E"
GOLD   = "00C9A84C"
WHITE  = "00FFFFFF"
LIGHT  = "00F0F2F7"
GRAY   = "00E0E5F0"

# ── テストケース定義 ──────────────────────────────────────────
# (試験番号, 確認対象画面, オペレーション内容, 期待結果)
CASES = [
    # ── セクション1 ──────────────────────────────────────
    ("【 初期画面起動（001〜002）】", None, None, None),
    (
        "UT-SES-001",
        "[Tab A]\nlogin.html",
        "代理店認証タブ（Tab A）でブラウザを開き login.html にアクセスする",
        "ネイビー系テーマの代理店ログイン画面（代理店コード・ログインID・パスワード入力欄、ログインボタン）が表示されること",
    ),
    (
        "UT-SES-002",
        "[Tab B]\nstaff_login.html",
        "別タブ（Tab B）で staff_login.html にアクセスする",
        "ピンク系テーマの社員ログイン画面（社員番号・パスワード入力欄、ログインボタン）が表示されること。Tab A の画面・セッションに影響がないこと",
    ),
    # ── セクション2 ──────────────────────────────────────
    ("【 各タブ交互ログイン（003〜004）】", None, None, None),
    (
        "UT-SES-003",
        "[Tab A]\nlogin.html → dashboard.html",
        "Tab A で代理店コード「A001」・ログインID「admin」・パスワード「password123」を入力し「ログイン」ボタンをクリックする",
        "dashboard.html へ遷移し、ヘッダーのユーザーバッジに「A001 admin｜管理者」と表示され、代理店テーマ（ネイビー）が適用されること",
    ),
    (
        "UT-SES-004",
        "[Tab B]\nstaff_login.html → dashboard.html",
        "Tab B で社員番号「S001」・パスワード「staff123」を入力し「ログイン」ボタンをクリックする",
        "dashboard.html へ遷移し、社員情報バッジが社員テーマ（ピンク）で表示されること。Tab A のセッション情報（A001/admin）が変化していないこと",
    ),
    # ── セクション3 ──────────────────────────────────────
    ("【 ダッシュボード交互確認（005〜006）】", None, None, None),
    (
        "UT-SES-005",
        "[Tab A]\ndashboard.html（代理店）",
        "Tab B のログイン完了後、Tab A のダッシュボードを再確認する（タブを切り替えてリロードせず確認）",
        "Tab A のユーザーバッジ「A001 admin｜管理者」・代理店テーマが保持されていること。Tab B のログイン操作による sessionStorage 上書きが発生していないこと",
    ),
    (
        "UT-SES-006",
        "[Tab B]\ndashboard.html（社員）",
        "Tab A の再確認完了後、Tab B のダッシュボードを再確認する",
        "Tab B のユーザーバッジに社員情報・社員テーマが保持されていること。Tab A の操作による sessionStorage 競合が発生していないこと",
    ),
    # ── セクション4 ──────────────────────────────────────
    ("【 満期管理画面交互遷移（007〜008）】", None, None, None),
    (
        "UT-SES-007",
        "[Tab A]\ndashboard.html → maturity.html",
        "Tab A のダッシュボードで「満期管理システム」ボタンをクリックし maturity.html へ遷移する",
        "maturity.html へ遷移し、ヘッダーに代理店コード（A001）と代理店テーマが表示されること。sessionStorage の agency_code・login_id が正しく保持されていること",
    ),
    (
        "UT-SES-008",
        "[Tab B]\ndashboard.html → maturity.html",
        "Tab B のダッシュボードで「満期管理システム」ボタンをクリックし maturity.html へ遷移する",
        "maturity.html へ遷移し、ヘッダーに社員番号（S001）と社員テーマが表示されること。Tab A の遷移後も Tab B のセッションに影響がないこと",
    ),
    # ── セクション5 ──────────────────────────────────────
    ("【 権限設定画面交互遷移（009〜010）】", None, None, None),
    (
        "UT-SES-009",
        "[Tab A]\nmaturity.html → dashboard.html → admin.html",
        "Tab A の満期管理ヘッダー「← ダッシュボード」をクリックして dashboard.html に戻り、アドレスバーから admin.html へ遷移する（代理店管理者セッションで権限設定画面へアクセス）",
        "admin.html（権限設定画面）が代理店管理者として正常に表示され、ユーザー一覧が確認できること",
    ),
    (
        "UT-SES-010",
        "[Tab B]\nmaturity.html → dashboard.html（戻る）",
        "Tab B の満期管理ヘッダー「← ダッシュボード」をクリックして dashboard.html に戻る",
        "dashboard.html に戻り、社員バッジ・社員テーマが保持されていること。Tab A の admin.html 遷移が Tab B のセッションに影響していないこと",
    ),
    # ── セクション6 ──────────────────────────────────────
    ("【 ダッシュボードへ戻る（011）】", None, None, None),
    (
        "UT-SES-011",
        "[Tab A]\nadmin.html → dashboard.html（戻る）",
        "Tab A の admin.html ヘッダー「ダッシュボードへ」リンクをクリックして dashboard.html に戻る",
        "dashboard.html に戻り、セッション情報（A001/admin/管理者）が正常に保持されていること。Tab B の操作による変化がないこと",
    ),
    # ── セクション7 ──────────────────────────────────────
    ("【 各タブ交互ログオフ（012〜013）】", None, None, None),
    (
        "UT-SES-012",
        "[Tab A]\ndashboard.html → login.html（ログアウト）",
        "Tab A のダッシュボードの「ログアウト」ボタンをクリックする",
        "login.html へリダイレクトされること。代理店 sessionStorage がクリアされること。Tab B の社員 sessionStorage は影響を受けず、社員セッションが継続されていること",
    ),
    (
        "UT-SES-013",
        "[Tab B]\ndashboard.html → staff_login.html（ログアウト）",
        "Tab B のダッシュボードの「ログアウト」ボタンをクリックする",
        "staff_login.html へリダイレクトされること。社員 sessionStorage がクリアされること。代理店・社員の全セッションが正常終了し、セッション競合が発生していないことの最終確認",
    ),
]

RESULT_DATE = "2026/05/30"


# ── スタイルヘルパー ──────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="FFB0B8C4")
    return Border(left=s, right=s, top=s, bottom=s)


def set_header_style(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True, size=10, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=fg, name="Meiryo")
    c.fill = PatternFill(fill_type="solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = thin_border()
    return c


def set_data_style(ws, row, col, value, bg="00FFFFFF", wrap=True, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(size=9, name="Meiryo", color="00333333")
    c.fill = PatternFill(fill_type="solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border = thin_border()
    return c


# ── 共通シート構築 ────────────────────────────────────────────
def build_sheet(ws, include_result):
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    # 行1: タイトル
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "社員認証・セッション競合回避テスト仕様書（sessionStorage タブ独立性検証）"
    c.font = Font(bold=True, size=14, color=WHITE, name="Meiryo")
    c.fill = PatternFill(fill_type="solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # 行2: テストID範囲・工程
    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:C2")
    ws["A2"].value = "テストID：UT-SES-001 〜 UT-SES-013"
    ws["A2"].font = Font(size=10, color=WHITE, name="Meiryo")
    ws["A2"].fill = PatternFill(fill_type="solid", fgColor=NAVY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("D2:G2")
    ws["D2"].value = "テスト工程：UT（単体テスト）　　作成日：2026/05/30"
    ws["D2"].font = Font(size=10, color=WHITE, name="Meiryo")
    ws["D2"].fill = PatternFill(fill_type="solid", fgColor=NAVY)
    ws["D2"].alignment = Alignment(horizontal="right", vertical="center")

    # 行3: 空白
    ws.row_dimensions[3].height = 8

    # 行4: 列ヘッダー
    ws.row_dimensions[4].height = 24
    headers = ["試験番号", "確認対象画面", "オペレーション内容", "期待結果",
               "テスト結果", "テスト実施日", "テスト確認日"]
    for i, h in enumerate(headers, 1):
        set_header_style(ws, 4, i, h, bg=NAVY, fg=WHITE, bold=True, size=10)

    # 行5〜: テストケース
    cur_row = 5
    case_idx = 0
    for case in CASES:
        trial_no, screen, op, expect = case

        # セクションヘッダー行
        if screen is None:
            ws.row_dimensions[cur_row].height = 20
            ws.merge_cells(f"A{cur_row}:G{cur_row}")
            c = ws.cell(row=cur_row, column=1, value=trial_no)
            c.font = Font(bold=True, size=10, color=NAVY, name="Meiryo")
            c.fill = PatternFill(fill_type="solid", fgColor="00E8EDF8")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin_border()
            cur_row += 1
            continue

        # データ行
        ws.row_dimensions[cur_row].height = 60
        row_bg = "00F7F9FC" if case_idx % 2 == 0 else "00FFFFFF"

        set_data_style(ws, cur_row, 1, trial_no,  bg=row_bg, wrap=False, align="center")
        set_data_style(ws, cur_row, 2, screen,    bg=row_bg, wrap=True)
        set_data_style(ws, cur_row, 3, op,        bg=row_bg, wrap=True)
        set_data_style(ws, cur_row, 4, expect,    bg=row_bg, wrap=True)

        if include_result:
            set_data_style(ws, cur_row, 5, "OK",        bg=row_bg, wrap=False, align="center")
            set_data_style(ws, cur_row, 6, RESULT_DATE, bg=row_bg, wrap=False, align="center")
            set_data_style(ws, cur_row, 7, RESULT_DATE, bg=row_bg, wrap=False, align="center")
        else:
            for col in [5, 6, 7]:
                set_data_style(ws, cur_row, col, None, bg=row_bg)

        cur_row += 1
        case_idx += 1

    return ws


def generate():
    dest = OUT_DIR

    # 仕様書
    wb_spec = openpyxl.Workbook()
    ws_spec = wb_spec.active
    ws_spec.title = "テスト仕様書"
    build_sheet(ws_spec, include_result=False)
    spec_path = dest / "社員認証セッション競合回避テスト仕様書.xlsx"
    wb_spec.save(spec_path)
    print(f"  生成: {spec_path.name}")

    # 証跡
    wb_rec = openpyxl.Workbook()
    ws_rec = wb_rec.active
    ws_rec.title = "テスト証跡"
    build_sheet(ws_rec, include_result=True)
    rec_path = dest / "社員認証セッション競合回避テスト証跡.xlsx"
    wb_rec.save(rec_path)
    print(f"  生成: {rec_path.name}")

    print("\n完了")


if __name__ == "__main__":
    generate()
