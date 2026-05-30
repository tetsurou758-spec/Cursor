"""
社員認証セッション競合回避テスト証跡.xlsx に UT-SES-001〜013 の
個別シートを追加し、UT-004 と同一フォーマットでスクリーンショットを
貼り付けるスクリプト
"""
import io, pathlib
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_PATH = pathlib.Path(__file__).parent / "社員認証セッション競合回避テスト証跡.xlsx"
SS_DIR     = pathlib.Path(__file__).parent / "screenshots"

IMG_W = 1400
IMG_H = 860

# ── テストケース定義 ──────────────────────────────────────────
# (シート名, 行1タイトル, 行2確認内容, 使用スクリーンショット)
CASES = [
    (
        "UT-SES-001",
        "UT-SES-001　[Tab A] 代理店ログイン画面 表示確認",
        "確認内容：代理店認証タブ（Tab A）で login.html を開き、ネイビー系テーマの代理店ログイン画面が表示されることを確認",
        "01_tab_a_agency_login.png",
    ),
    (
        "UT-SES-002",
        "UT-SES-002　[Tab B] 社員ログイン画面 表示確認",
        "確認内容：別タブ（Tab B）で staff_login.html を開き、ピンク系テーマの社員ログイン画面が表示されること。Tab A のセッションに影響がないことを確認",
        "02_tab_b_staff_login.png",
    ),
    (
        "UT-SES-003",
        "UT-SES-003　[Tab A] 代理店ログイン → ダッシュボード遷移確認",
        "確認内容：Tab A で A001/admin/password123 でログインし dashboard.html へ遷移。「A001 admin｜管理者」バッジ・代理店テーマが表示されることを確認",
        "03_tab_a_dashboard_agency.png",
    ),
    (
        "UT-SES-004",
        "UT-SES-004　[Tab B] 社員ログイン → ダッシュボード遷移確認",
        "確認内容：Tab B で S001/staff123 でログインし dashboard.html へ遷移。社員バッジ・社員テーマ（ピンク）が表示されること。Tab A のセッションが変化していないことを確認",
        "04_tab_b_dashboard_staff.png",
    ),
    (
        "UT-SES-005",
        "UT-SES-005　[Tab A] Tab B ログイン後 ダッシュボード再確認（競合なし）",
        "確認内容：Tab B のログイン完了後に Tab A のダッシュボードを確認。ユーザーバッジ「A001 admin｜管理者」・代理店テーマが保持され sessionStorage 競合がないことを確認",
        "05_tab_a_dashboard_recheck.png",
    ),
    (
        "UT-SES-006",
        "UT-SES-006　[Tab B] Tab A 再確認後 ダッシュボード再確認（競合なし）",
        "確認内容：Tab A の再確認操作後に Tab B のダッシュボードを確認。社員バッジ・社員テーマが保持され sessionStorage 競合がないことを確認",
        "06_tab_b_dashboard_recheck.png",
    ),
    (
        "UT-SES-007",
        "UT-SES-007　[Tab A] 満期管理画面 遷移確認",
        "確認内容：Tab A のダッシュボードで「満期管理システム」ボタンをクリックし maturity.html へ遷移。ヘッダーに A001 と代理店テーマが表示されることを確認",
        "07_tab_a_maturity.png",
    ),
    (
        "UT-SES-008",
        "UT-SES-008　[Tab B] 満期管理画面 遷移確認（Tab A 遷移後の競合なし）",
        "確認内容：Tab B のダッシュボードで「満期管理システム」ボタンをクリックし maturity.html へ遷移。S001 と社員テーマが表示され Tab A の遷移による影響がないことを確認",
        "08_tab_b_maturity.png",
    ),
    (
        "UT-SES-009",
        "UT-SES-009　[Tab A] 権限設定画面(admin.html) 遷移確認",
        "確認内容：Tab A の満期管理から「← ダッシュボード」→ dashboard.html に戻り admin.html へ遷移。代理店管理者セッション（A001/admin）で権限設定画面が正常表示されることを確認",
        "09_tab_a_admin.png",
    ),
    (
        "UT-SES-010",
        "UT-SES-010　[Tab B] 満期管理→ダッシュボード戻り確認（Tab A admin.html 遷移後）",
        "確認内容：Tab B の満期管理から「← ダッシュボード」で dashboard.html に戻る。Tab A の admin.html 遷移後も社員セッション・社員テーマが保持されていることを確認",
        "10_tab_b_dashboard_back_from_maturity.png",
    ),
    (
        "UT-SES-011",
        "UT-SES-011　[Tab A] admin.html → ダッシュボード戻り確認（セッション保持）",
        "確認内容：Tab A の admin.html 「ダッシュボードへ」リンクで dashboard.html に戻る。セッション情報（A001/admin/管理者）が保持され Tab B 操作による変化がないことを確認",
        "11_tab_a_dashboard_back_from_admin.png",
    ),
    (
        "UT-SES-012",
        "UT-SES-012　[Tab A] 代理店ログアウト確認（Tab B セッション継続）",
        "確認内容：Tab A でログアウトボタンをクリックし login.html へリダイレクト。代理店 sessionStorage がクリアされ Tab B の社員セッションが継続中であることを確認",
        "12_tab_a_after_logout.png",
    ),
    (
        "UT-SES-013",
        "UT-SES-013　[Tab B] 社員ログアウト確認（全セッション正常終了）",
        "確認内容：Tab B でログアウトボタンをクリックし staff_login.html へリダイレクト。社員 sessionStorage がクリアされ全セッションが正常終了。セッション競合なしを最終確認",
        "13_tab_b_after_logout.png",
    ),
]

# ── スタイル定数（UT-004 と同一）──────────────────────────────
NAVY     = "000D2B5E"
GOLD_BG  = "00FFF8E6"
WHITE    = "00FFFFFF"
GRAY_TXT = "005A6480"
NAVY_TXT = "000D2B5E"
EXEC_INFO = "テスト実施日：2026/05/30　　環境：ローカルファイル（file://）　　ブラウザ：Chromium (Playwright headless=False)"


def resize_to_buf(img_path: pathlib.Path, w: int, h: int) -> io.BytesIO:
    with PILImage.open(img_path) as im:
        im = im.convert("RGB")
        im = im.resize((w, h), PILImage.LANCZOS)
        buf = io.BytesIO()
        im.save(buf, format="PNG", optimize=True)
        buf.seek(0)
    return buf


def build_case_sheet(wb, sheet_name, row1_title, row2_content, img_path):
    ws = wb.create_sheet(sheet_name)

    ws.column_dimensions["A"].width = 18.0

    # 行1: タイトル（navy背景・白太字）
    ws.row_dimensions[1].height = 28.0
    ws.merge_cells("A1:H1")
    c1 = ws["A1"]
    c1.value = row1_title
    c1.font  = Font(bold=True, size=12, color=WHITE, name="Meiryo")
    c1.fill  = PatternFill(fill_type="solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center")

    # 行2: 確認内容（薄黄背景・紺文字）
    ws.row_dimensions[2].height = 20.0
    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = row2_content
    c2.font  = Font(bold=False, size=10, color=NAVY_TXT, name="Meiryo")
    c2.fill  = PatternFill(fill_type="solid", fgColor=GOLD_BG)
    c2.alignment = Alignment(horizontal="left", vertical="center")

    # 行3: 実施情報（グレー文字）
    ws.row_dimensions[3].height = 16.0
    ws.merge_cells("A3:H3")
    c3 = ws["A3"]
    c3.value = EXEC_INFO
    c3.font  = Font(bold=False, size=9, color=GRAY_TXT, name="Meiryo")
    c3.alignment = Alignment(horizontal="left", vertical="center")

    # 行5〜: スクリーンショット
    buf    = resize_to_buf(img_path, IMG_W, IMG_H)
    xl_img = XLImage(buf)
    xl_img.width  = IMG_W
    xl_img.height = IMG_H
    ws.add_image(xl_img, "A5")

    return ws


def main():
    wb = load_workbook(EXCEL_PATH)

    # 既存の個別ケースシートを削除（再実行対応）
    for case_id, *_ in CASES:
        if case_id in wb.sheetnames:
            del wb[case_id]

    for sheet_name, row1, row2, img_name in CASES:
        img_path = SS_DIR / img_name
        if not img_path.exists():
            print(f"  [SKIP] {img_name} が見つかりません → _capture_screenshots.py を先に実行してください")
            continue
        build_case_sheet(wb, sheet_name, row1, row2, img_path)
        print(f"  [{sheet_name}] ← {img_name}")

    wb.save(EXCEL_PATH)
    print(f"\n保存完了: {EXCEL_PATH}")
    print(f"シート構成: {wb.sheetnames}")


if __name__ == "__main__":
    main()
